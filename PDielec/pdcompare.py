#!/usr/bin/env python
#
# Copyright 2015 John Kendrick
#
# This file is part of PDielec
#
# This program is free software; you can redistribute it and/or modify
# it under the terms of the MIT License
#
# This program is distributed in the hope that it will be useful,
# but WITHOUT ANY WARRANTY; without even the implied warranty of
# MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.
#
# You should have received a copy of the MIT License
# along with this program, if not see https://opensource.org/licenses/MIT
#
'''PDielec driver program to calculate dielectric response at infrared and THz frequencies'''
from __future__ import print_function
import sys
import numpy as np
from openpyxl import load_workbook


def main():
    '''Main Driver routine for pdcompare - reads a collection of spreadsheets with spectra and compares the spectra'''
    def show_usage():
        '''Show pdcompare usage'''
        print('pdcompare file1 file2...')
        print('pdcompare: Calculates the cross-correlation of spectra calculated using pdielec')
        print('         -column column  Take the data for cross correlation from column')
        print('                         C Averaged')
        print('                         D MG      ')
        print('                         E Mie/0.1 ')
        print('                         F Mie/1.0 ')
        print('                         G Mie/2.0 ')
        print('                         H Mie/3.0 ')
        print('         -rmax rmax  Use rows from rmin to rmax              ')
        print('         -rmin rmin  Use rows from rmin to rmax (rows start from 2)')
        print('         -sheet [molar/absorption/real/imaginary/atr]')
        print('         -excel filename')
        return

    # check usage
    if len(sys.argv) <= 1:
        show_usage()
        exit()

    sheet_dict = {
        "molar"       : "Molar Absorption",
        "absorption"  : "Absorption",
        "real"        : "Real Permittivity",
        "imaginary"   : "Imaginary Permittivity",
        "atr"         : "ATR Reflectance",
    }
    # Begin processing of command line
    tokens = sys.argv[1:]
    ntokens = len(tokens)
    itoken = 0
    names = []
    rmin = 0
    rmax = 0
    column = 'D'
    sheet = 'molar'
    while itoken < ntokens:
        token = tokens[itoken]
        if token == '-rmin':
            itoken += 1
            rmin = int(tokens[itoken])
        elif token == '-rmax':
            itoken += 1
            rmax = int(tokens[itoken])
        elif token == '-excel':
            itoken += 1
            excelfile = tokens[itoken]
        elif token == '-column':
            itoken += 1
            column = tokens[itoken]
        elif token == '-sheet':
            itoken += 1
            sheet= tokens[itoken]
        else:
            names.append(token)
        itoken += 1
        # end loop over tokens

    if len(names) <= 0:
        print('No files were specified')
        show_usage()
        exit(1)

    print('Comparison based on ', sheet, sheet_dict[sheet])
    sheet = sheet_dict[sheet]
    print('Comparision of spectra based on column: ',column)
    if excelfile != '':
        print('Output will be sent to the excel file: ',excelfile)
    size = len(names)
    columns = []
    lags = np.zeros( (size,size ) )
    correlations = np.eye( size )
    # Use the first file name to define the frequency range
    # and the range of rows to be treated
    wb1 = load_workbook(filename=names[0], read_only=True)
    ws1 = wb1[sheet]
    max_rows1 = ws1.max_row
    max_cols1 = ws1.max_column
    print('Maximum number of rows',max_rows1)
    print('Maximum number of cols',max_cols1)
    # rmax and rmin are set by the first spread sheet
    if rmin == 0:
        rmin = 2
    if rmax == 0:
        rmax = max_rows1
    range1 = '{}{}'.format('B',rmin)
    range2 = '{}{}'.format('B',rmax)
    frequencies = np.array([[i.value for i in j] for j in ws1[range1:range2]])
    frequencies = frequencies[:,0]
    freq_scale = (frequencies[1] - frequencies[0])
    print('Frequencies',frequencies)
    print('Frequency scale',freq_scale)
    # Go through the file names and store the required column of numbers
    for f1_name in names:
        print('Loading work book ', f1_name)
        wb1 = load_workbook(filename=f1_name, read_only=True)
        # print('Work sheet names for ',f1_name)
        # print(wb1.get_sheet_names())
        ws1 = wb1[sheet]
        range1 = '{}{}'.format(column,rmin)
        range2 = '{}{}'.format(column,rmax)
        col1 = np.array([[i.value for i in j] for j in ws1[range1:range2]])
        # Convert to a 1D array
        col1 = col1[:,0]
        # Store the normalised signal
        col1 = ( col1 - np.mean(col1)) / ( np.std(col1) * np.sqrt(len(col1)) )
        columns.append(col1)
        # print(columns[-1])
    # Print the new row min and max
    print('Final rmin is ', rmin)
    print('Final rmax is ', rmax)
    for i,col1 in enumerate(columns):
        for j,col2 in enumerate(columns):
            if i > j:
                #print('Correlation',i,j)
                # Calculate correlation using same and full seem to produce the same results
                correlation = np.correlate(col1,col2,mode='full')
                lag = np.argmax(correlation) - (len(correlation)-1)/2
                #print('Old lag',lag,len(correlation))
                lags[i,j] = lag
                lags[j,i] = lags[i,j]
                correlations[i,j] = np.max(correlation)
                correlations[j,i] = correlations[i,j]
                #print(correlation)
                #print('Lag = ',lags[i,j])
                #print('Maximum = ',correlations[i,j])
        # end for j,f2
    # end for i, f1
    print('Lags (steps)')
    print(lags)
    lags = freq_scale*lags
    print('Lags (cm-1)')
    print(lags)
    print('correlations')
    print(correlations)
    if excelfile != "":
        import xlsxwriter as xlsx
        workbook = xlsx.Workbook(excelfile)
        worksheet = workbook.add_worksheet('Lags_cm1')
        for i,name1 in enumerate(names):
           worksheet.write(0,i+1,name1)
           worksheet.write(i+1,0,name1)
           for j,name2 in enumerate(names):
               worksheet.write(j+1,i+1,lags[j,i])
               worksheet.write(i+1,j+1,lags[i,j])
        worksheet = workbook.add_worksheet('Correlations')
        for i,name1 in enumerate(names):
           worksheet.write(0,i+1,name1)
           worksheet.write(i+1,0,name1)
           for j,name2 in enumerate(names):
               worksheet.write(j+1,i+1,correlations[j,i])
               worksheet.write(i+1,j+1,correlations[i,j])
        print('Finished write the spread sheet to ',excelfile)
        workbook.close()


if __name__ == '__main__':
    main()
