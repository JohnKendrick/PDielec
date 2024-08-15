#
# Copyright 2024 John Kendrick & Andrew Burnett
#
# This file is part of PDielec
#
# This program is free software; you can redistribute it and/or modify
# it under the terms of the MIT License
#
# This program is distributed in the hope that it will be useful,
# but WITHOUT ANY WARRANTY; without even the implied warranty of
# MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.
#
# You should have received a copy of the MIT License along with this program, if not see https://opensource.org/licenses/MIT
#
"""FitterTab module."""
import os.path

# Import plotting requirements
import matplotlib
import matplotlib.figure
import numpy as np
from matplotlib.backends.backend_qtagg import FigureCanvasQTAgg as FigureCanvas
from matplotlib.backends.backend_qtagg import NavigationToolbar2QT as NavigationToolbar
from qtpy.QtCore import QCoreApplication, Qt
from qtpy.QtWidgets import (
    QCheckBox,
    QComboBox,
    QDoubleSpinBox,
    QFileDialog,
    QFormLayout,
    QHBoxLayout,
    QLabel,
    QLineEdit,
    QProgressBar,
    QPushButton,
    QSizePolicy,
    QSpinBox,
    QTableWidgetItem,
    QTabWidget,
    QVBoxLayout,
    QWidget,
)
from scipy.interpolate import interp1d
from scipy.optimize import minimize

from PDielec import Calculator
from PDielec.GUI.SettingsTab import FixedQTableWidget
from PDielec.Utilities import Debug


def is_float(element):
    """Test to see if element is a float.

    Parameters
    ----------
    element : str
        A string to be analysed to see if it is a float

    Returns
    -------
    bool
        True if the element is a float, False otherwise.

    """
    #If you expect None to be passed:
    if element is None: 
        return False
    try:
        float(element)
        return True
    except ValueError:
        return False

def find_delimiter(filename):
    """Find the delimiter used in a CSV file.

    Parameters
    ----------
    filename : str
        The path to the file for which the delimiter is to be determined.

    Returns
    -------
    str
        The detected delimiter character.

    Notes
    -----
    This function uses the `csv.Sniffer` class to deduce the delimiter of
    the given file by analyzing the first 5000 bytes of the file.

    """        
    import csv
    sniffer = csv.Sniffer()
    with open(filename) as fp:
        return sniffer.sniff(fp.read(5000)).delimiter

def read_experimental_file(file_name,frequency_column=1, spectrum_column=2, sheet_name=None, debug=False):
    """Read an experimental file.

    Parameters
    ----------
    file_name : str
        An xlsx or a csv file containing the experimental spectrum
    frequency_column : int
        The frequency column in the spread sheet, default is 1.  The first column in a sheet is column 1
    spectrum_column : int
        The spectrum column in the spread sheet, default is 2.  The first column in a sheet is column 1
    sheet_name : str
        The name of the spreadsheet to be used to extract the frequency and spectrum, default is None
    debug : boolean
        True if debug information is needed, default is False

    Returns
    -------
    experimental_frequencies : an array of floats
        The experimental frequencies in cm-1
    experimental_spectrum : an array of floats
        The experimental spectrum

    Errors
    ------
    If any errors occur, such as a file not existing, empty arrays/lists are returned, :w

    """ 
    experimental_frequencies = []
    experimental_spectrum = []
    if debug:
        print("FitterTab: Start:: read_experimental_file",file_name)
    if not os.path.isfile(file_name):
        if debug:
            print("FitterTab: Finished:: read_experimental_file does not exist",file_name)
        return experimental_frequencies, experimental_spectrum
    if file_name.endswith(".csv"):
        # Read in a csv file, discard alphanumerics
        import csv
        delimiter = find_delimiter(file_name)
        with open(file_name) as fd:
            if debug:
                print("FitterTab: Reading csv file::")
            csv_reader = csv.reader(fd, delimiter=delimiter)
            for row in csv_reader:
               # attempt to convert string to float
               if is_float(row[frequency_column-1]) and is_float(row[spectrum_column-1]):
                   experimental_frequencies.append(float(row[frequency_column-1]))
                   experimental_spectrum.append(float(row[spectrum_column-1]))
    else:
        # Read in a xlsx file, discard alphanumerics
        import warnings

        from openpyxl import load_workbook
        if debug:
            print("FitterTab: Reading xlsx file::")
        with warnings.catch_warnings():
            warnings.simplefilter("ignore")
            wb = load_workbook(filename=file_name, read_only=True)
        sheetnames = wb.sheetnames
        ws = wb[sheet_name] if sheet_name in sheetnames else wb.worksheets[0]
        for row in ws.rows:
            if isinstance(row[frequency_column-1].value, (int, float, complex)):
                experimental_frequencies.append(row[frequency_column-1].value)
                experimental_spectrum.append(row[spectrum_column-1].value)
        wb.close()
    if debug:
        print("FitterTab: Finished:: read_experimental_file")
    return experimental_frequencies, experimental_spectrum

def resample_experimental_spectrum(calculated_frequencies,
                                   experimental_frequencies, experimental_spectrum,
                                   baseline_removal="False", HPFilter_lambda=7.0,
                                   debug=False, ):
    """Resample the experimental spectrum to match calculated frequencies and optionally apply baseline removal.

    This method modifies the experimental spectrum to align with the calculated 
    frequencies. It pads the spectrum at both ends with zeros outside the calculated
    frequencies range, then uses cubic interpolation over the whole range. 
    If 'baseline_removal' is enabled, applies the Hodrick-Prescott filter 
    for baseline removal on the resampled spectrum.

    Parameters
    ----------
    calculated_frequencies : array of floats
        The frequencies the caculated spectrum is tabulated 
    experimental_frequencies : array of floats
        The frequencies the experimental spectrum is tabulated 
    experimental_spectrum : array of floats
        The experimental spectrum
    baseline_removal : boolean
        Optional switches on baseline removal using a Hodrick-Prescott filter, default is False
    HPFilter_lambda : float
        The value of lambda in in the Hodrick-Prescott filter, default value is 7.0
    debug : boolean
        Optional switches on some debug information, default is False

    Returns
    -------
    None
        Modifies the instance's resampled_experimental_spectrum in place.

    """        
    if (debug):
        print("Start:: Resample_experimental_spectrum")
    #
    # If the experimental frequencies starts at a higher frequency 
    # than the calculated frequencies then add new frequencies to pad the range out
    #
    xaxis = calculated_frequencies
    # indices is true if the experimental frequency is outside the range of the calculated
    indices = xaxis < experimental_frequencies[0]
    
    padded_xaxis = xaxis[indices]
    padded_yaxis = np.array([ 0 for index in indices if index ])
    # Add the experimental frequencies
    padded_xaxis = np.append(padded_xaxis,experimental_frequencies)
    padded_yaxis = np.append(padded_yaxis,experimental_spectrum)
    # If the experimental frequencies ends at a lower frequency 
    # than the calculated frequencies then add new frequencies to pad the range out
    indices = xaxis > experimental_frequencies[-1]
    padded_xaxis = np.append(padded_xaxis,xaxis[indices])
    padded_yaxis = np.append(padded_yaxis, np.array([ 0 for index in indices if index ]) )
    # 
    # Create a function using the padded frequencies to calculate the spectrum at the calculated frequencies
    #
    f = interp1d(padded_xaxis, padded_yaxis, kind="cubic",fill_value="extrapolate")
    # Store the experimental spectrum at the calculated frequencies
    resampled_experimental_spectrum = f(xaxis)
    if baseline_removal:
        # Apply a Hodrick-Prescott filter to remove the background
        resampled_experimental_spectrum = Calculator.hodrick_prescott_filter(
                                      resampled_experimental_spectrum, 0.01,
                                      HPFilter_lambda, 10)
    if (debug):
        print("Finished:: Resample_experimental_spectrum")
    return resampled_experimental_spectrum

def calculateCrossCorrelation(xaxis, calculated_spectrum, experimental_spectrum,
                              scaling_factor=1.0, debug=False):
    """Calculate the cross-correlation between the experimental spectrum and a calculated spectrum, with an applied scaling factor.

    The calculated and experimental spectra share the same frequencies (xaxis).  If a scaling factor is applied, 
    the calculated spectrum x-axis is rescaled according to the scaling factor.

    Parameters
    ----------
    xaxis : 1d array of floats
        An array frequencies valid for both the calculated and experimental spectra
    calculated_spectrum : 1d array of floats
        The calculated spectrum
    experimental_spectrum : 1d array of floats
        The experimental spectrum
    scaling_factor : float, optional
        The factor by which the x-axis of the calculated spectrum is scaled, default is 1.0
    debug : boolean, optional
        Set to true for debugging information

    Returns
    -------
    tuple
        A tuple containing:
        - lag: The lag at which the maximum correlation occurs.
        - max_corr: The maximum correlation value.
        - center_corr: The correlation value at the center (zero lag).

    Notes
    -----
    The calculated spectrum's x-axis is rescaled according to the `scaling_factor`. The correlation is performed on the normalized (zero-mean unit-variance) spectra. Interpolation of the calculated spectrum onto the experimental spectrum's x-axis is performed using cubic interpolation.

    """        
    if debug:
        print("Start:: calculateCrossCorrelation",scaling_factor)
    # Calculate the cross correlation coefficient between the experimental and the first scenario
    if len(experimental_spectrum) == 0 or len(calculated_spectrum) == 0 or len(xaxis) == 0:
        if debug:
            print("The experimental spectrum has not been specified")
            if len(experimental_spectrum) == 0:
                print("calculateCrossCorrelation experimental_spectrum is not defined")
            if len(calculated_spectrum) == 0:
                print("calculateCrossCorrelation experimental_spectrum is not defined")
            if len(xaxis) == 0:
                print("calculateCrossCorrelation the x-axis frequencies are not defined")
            print("Finshed:: calculateCrossCorrelation",scaling_factor)
        return (0.0,0.0,0.0)
    # col1 contains the experimental spectrum
    col1 = np.array(experimental_spectrum)
    col1 = ( col1 - np.mean(col1)) / ( np.std(col1) * np.sqrt(len(col1)) )
    # col2 contains the calculated spectrum
    col2 = np.array(calculated_spectrum)
    # The new xaxis for the calculated spectrum is scaling_factor*xaxis
    f = interp1d(scaling_factor*np.array(xaxis), col2, kind="cubic",fill_value="extrapolate")
    col2 = f(xaxis)
    col2 = ( col2 - np.mean(col2)) / ( np.std(col2) * np.sqrt(len(col2)) )
    correlation = np.correlate(col1, col2,  mode="full")
    lag = np.argmax(correlation) - (len(correlation)-1)/2
    lag = (xaxis[1] - xaxis[0]) * lag
    if debug:
        print("lag , max(corr), index", lag,np.max(correlation),correlation[int((len(correlation)-1)/2)])
        print("Finshed:: calculateCrossCorrelation",scaling_factor)
    return (lag,np.max(correlation),correlation[int((len(correlation)-1)/2)])

def calculateSpectralDifference(xaxis,calculated_spectrum,experimental_spectrum,
                            spectral_threshold=0.05, scaling_factor=1.0, debug=False):
    """Calculate the root mean square error (RMSE) between a renormalised experimental spectrum and calculated spectrum after applying a scaling factor.

    Parameters
    ----------
    xaxis : 1d array of floats
        An array frequencies valid for both the calculated and experimental spectra
    calculated_spectrum : 1d array of floats
        The calculated spectrum
    experimental_spectrum : 1d array of floats
        The experimental spectrum
    scaling_factor : float, optional
        The factor by which the x-axis of the calculated spectrum is scaled, default is 1.0
    spectral_threshold : float,optional
        Only signals greater than the threshold are considered in the difference calculation. Default is 0.05
    debug : boolean, optional
        Set to true for debugging information

    Returns
    -------
    float
        The root mean square error between the resampled experimental spectrum and the scaled calculated spectrum.

    Notes
    -----
    - This method modifies the spectra by scaling, normalization, and thresholding before calculating the RMSE.
    - If the resampled experimental spectrum is empty, the method returns 0.0 immediately.
    - Uses cubic interpolation for scaling the calculated spectrum.
    - Normalization is performed based on the maximum value in the experimental spectrum.

    """        
    if debug:
        print("FitterTab: calculateSpectralDifference",scaling_factor)
    # Calculate the spectral difference  between the experimental and the first scenario
    if len(experimental_spectrum) == 0:
        return 0.0
    # col1 contains the experimental spectrum
    col1 = np.array(experimental_spectrum)
    maxcol1 = np.max(col1)
    col1 = col1 / maxcol1
    col1[ col1< spectral_threshold ] = 0.0
    # col2 contains the calculated spectrum
    col2 = np.array(calculated_spectrum)
    # The new xaxis for the calculated spectrum is scaling_factor*xaxis
    f = interp1d(scaling_factor*np.array(xaxis), col2, kind="cubic",fill_value="extrapolate")
    col2 = f(xaxis)
    # Note the normalisation of col2 is to col1 (the experimental spectrum)
    col2 = col2 / maxcol1
    col2[ col2< spectral_threshold ] = 0.0
    diff = col1 - col2
    rmse = np.sqrt(np.dot(diff,diff)/len(col2))
    if debug:
        print("FitterTab: rmse",rmse)
        print("FitterTab: Finished:: optimiseFunction")
    return rmse

class FitterTab(QWidget):
    """Initialize the main FitterTab with its components and connect signals to slots.

    A class for managing and displaying a spectroscopic fitting interface within a QWidget.
    This constructor sets up the UI for a spectral analysis widget, initializing all the GUI components such as buttons, combo boxes, spin boxes, and plotting canvases. It also sets up the initial state for various attributes related to the spectral analysis settings and data.

    Parameters
    ----------
    parent : QWidget
        The parent widget.
    debug : bool, optional
        Flag to indicate whether debugging messages should be printed, by default False.

    Attributes
    ----------
    parent : QWidget
        The parent QWidget within which this `FitterTab` is contained.
    debug : bool, optional
        A flag to activate debug mode which provides additional output for development and troubleshooting. Defaults to False.

    Notes
    -----
    This class is designed to provide a user interface for spectral fitting tasks, allowing for adjustments of various fitting parameters and visualization of fitting results. It leverages Qt widgets for the graphical user interface components and matplotlib for plotting.

    The class initializes with customizable settings related to spectral fitting including file names, fitting types, iterations, scaling factors, and baseline correction options. It supports reading experimental data from .xlsx or .csv files, setting fitting parameters through a dynamically generated form, and presenting the fitting results graphically with options to adjust the view.

    Methods
    -------
    __init__(self, parent, debug=False)
        Constructor to initialize the spectroscopic fitting interface with the given parent and debug settings.
    on_iterations_sb_changed(self)
        Handle changes to the number of iterations spinbox value.
    on_independent_yaxes_cb_changed(self, value)
        Handle changes to the independent y-axes checkbox state.
    on_optimise_frequency_scaling_cb_changed(self, value)
        Handle changes to the optimise frequency scaling checkbox state.
    on_spectrum_scaling_cb_changed(self, value)
        Handle changes to the spectrum scaling checkbox state.
    on_hpfilter_lambda_sb_changed(self, value)
        Handle changes to the HP filter lambda spinbox value.
    on_spectrum_scaling_factor_sb_changed(self, value)
        Handle changes to the spectrum scaling factor spinbox value.
    on_baseline_cb_changed(self, value)
        Handle changes to the baseline correction checkbox state.
    on_spectral_difference_threshold_sb_changed(self, value)
        Handle changes to the spectral difference threshold spinbox value.
    on_frequency_scaling_factor_sb_changed(self, value)
        Handle changes to the frequency scaling factor spinbox value.
    replotButton1Clicked(self)
        Handle the event when the first replot button is clicked.
    replotButton2Clicked(self)
        Handle the event when the second replot button with frequency shift is clicked.
    plot(self, experiment, xs, ys, legends, label)
        Plots the fitting results with optional experiment data overlay.
    fittingButtonClicked(self)
        Initiates the fitting process when the fitting button is clicked.
    optimiseFit(self)
        Optimises the fitting parameters.
    optimiseFunction(self, variables)
        Function to be optimised during the fitting process.
    redraw_sigmas_tw(self)
        Redraws the table widget for sigma values based on current settings.
    on_fitting_type_cb_activated(self, index)
        Handle changes to the fitting type combo box selection.
    on_scenario_cb_activated(self, index)
        Handle selection changes in the scenario combo box.
    on_spectrafile_le_return(self)
        Handle actions to take when the spectra file line edit has a return event.
    on_spectrafile_le_changed(self, text)
        Handle changes to the spectra file line edit text.
    on_sigmas_tw_itemChanged(self, item)
        Responds to item changes in the sigmas table widget.
    print_settings(self)
        Prints the current settings to the debugger.
    replot(self)
        Handle the replotting of spectra.
    refresh(self, force=False)
        Refreshes the interface based on current settings and data.
    requestRefresh(self)
        Requests a refresh of the interface.

    """

    def __init__(self, parent, debug=False):
        """Initialize the main widget with its components and connect signals to slots.

        This constructor sets up the UI for a spectral analysis widget, initializing all the GUI components such as buttons, combo boxes, spin boxes, and plotting canvases. It also sets up the initial state for various attributes related to the spectral analysis settings and data.

        Parameters
        ----------
        parent : QWidget
            The parent widget.
        debug : bool, optional
            Flag to indicate whether debugging messages should be printed, by default False.

        Notes
        -----
        The function initializes a complex UI for spectral analysis, handling settings such as file names for experimental data, fitting types, frequency scaling, and more. It also prepares the UI for user interactions, setting up signals and slots for UI components like buttons, line edits, and combo boxes. Additionally, it initializes data structures for storing spectral data and analysis results.

        """        
        super(QWidget, self).__init__(parent)
        global debugger
        debugger = Debug(debug,"FitterTab:")
        debugger.print("Start:: Initialising")
        self.refreshRequired = True
        self.calculationInProgress = False
        self.settings = {}
        self.notebook = parent
        self.settings["Experimental file name"] = ""
        self.settings["Plot title"] = "Experimental and Calculated Spectral Comparison"
        self.settings["Fitting type"] = "Minimise x-correlation"
        self.fitting_type_definitions = ["Minimise x-correlation", "Minimise spectral difference"]
        self.settings["Number of iterations"] = 20
        self.settings["Frequency scaling factor"] = 1.0
        self.settings["Optimise frequency scaling"] = False
        self.settings["Spectrum scaling"] = False
        self.settings["Spectrum scaling factor"] = 1.0
        self.settings["Independent y-axes"] = True
        self.settings["Spectral difference threshold"] = 0.05
        self.settings["HPFilter lambda"] = 7.0
        self.settings["Baseline removal"] = False
        self.settings["Scenario index"] = len(self.notebook.scenarios) - 1
        self.scenario_legends = [ scenario.settings["Legend"] for scenario in self.notebook.scenarios ]
        self.lastButtonPressed = self.replotButton1Clicked
        self.plot_frequency_shift = False
        self.xcorr0=0.0
        self.xcorr1=0.0
        self.lag=0.0
        self.experimental_file_has_been_read = False
        self.sigmas_cm1 = []
        self.modes_fitted = []
        self.modes_selected = []
        self.experimental_frequencies = []
        self.experimental_spectrum = []
        self.resampled_experimental_spectrum = []
        self.frequency_units = "wavenumber"
        # Create a tab
        vbox = QVBoxLayout()
        form = QFormLayout()
        #
        # Ask for the experimental spread sheet
        #
        self.spectrafile_le = QLineEdit(self)
        self.spectrafile_le.setToolTip("Provide the name of an .xlsx or a .csv file containing the experimental spectrum.\nThe spreadsheet should have two columns, with at most a one line heading.\nThe first column contains the frequencies in cm-1.  The second contains the experimental spectrum.")
        self.spectrafile_le.setText(self.settings["Experimental file name"])
        self.spectrafile_le.returnPressed.connect(self.on_spectrafile_le_return)
        self.spectrafile_le.textChanged.connect(self.on_spectrafile_le_changed)
        #
        # Select the type of fitting we are going to use
        #
        self.fitting_type_cb = QComboBox(self)
        self.fitting_type_cb.setToolTip("What type of fitting?")
        self.fitting_type_cb.addItems(self.fitting_type_definitions)
        self.fitting_type_cb.activated.connect(self.on_fitting_type_cb_activated)
        self.fitting_type_cb.setCurrentIndex(self.fitting_type_definitions.index(self.settings["Fitting type"]))
        #
        # Select the type of plot we are going to use
        #
        self.scenario_cb = QComboBox(self)
        self.scenario_cb.setToolTip("Which scenario will be used in the fit?")
        self.scenario_cb.addItems(self.scenario_legends)
        self.scenario_cb.activated.connect(self.on_scenario_cb_activated)
        self.scenario_cb.setCurrentIndex(self.settings["Scenario index"])
        #
        # See if we want frequency scaling
        #
        self.frequency_scaling_factor_sb = QDoubleSpinBox(self)
        self.frequency_scaling_factor_sb.setToolTip("Frequency scaling factor")
        self.frequency_scaling_factor_sb.setRange(0.000001,10000000.0)
        self.frequency_scaling_factor_sb.setDecimals(4)
        self.frequency_scaling_factor_sb.setSingleStep(0.1)
        self.frequency_scaling_factor_sb.setValue(self.settings["Frequency scaling factor"])
        self.frequency_scaling_factor_sb.setToolTip("Set the value for scaling the frequency axis of the calculated spectrum")
        self.frequency_scaling_factor_sb.valueChanged.connect(self.on_frequency_scaling_factor_sb_changed)
        #
        # See if we want base line removal
        #
        self.baseline_cb = QCheckBox(self)
        self.baseline_cb.setToolTip("Apply base line correction")
        self.baseline_cb.setText("")
        self.baseline_cb.setLayoutDirection(Qt.RightToLeft)
        if self.settings["Baseline removal"]:
            self.baseline_cb.setCheckState(Qt.Checked)
        else:
            self.baseline_cb.setCheckState(Qt.Unchecked)
        self.baseline_cb.stateChanged.connect(self.on_baseline_cb_changed)

        # Hodrick-Prescott Filter Lambda
        self.hpfilter_lambda_sb = QDoubleSpinBox(self)
        self.hpfilter_lambda_sb.setRange(0.0,10000000.0)
        self.hpfilter_lambda_sb.setDecimals(1)
        self.hpfilter_lambda_sb.setSingleStep(1)
        self.hpfilter_lambda_sb.setValue(self.settings["HPFilter lambda"])
        self.hpfilter_lambda_sb.setToolTip("The Hodrick-Prescott filter lambda value")
        self.hpfilter_lambda_sb.valueChanged.connect(self.on_hpfilter_lambda_sb_changed)
        # Spectral difference threshold
        self.spectral_difference_threshold_sb = QDoubleSpinBox(self)
        self.spectral_difference_threshold_sb.setRange(0.000001,10.0)
        self.spectral_difference_threshold_sb.setDecimals(2)
        self.spectral_difference_threshold_sb.setSingleStep(0.1)
        self.spectral_difference_threshold_sb.setValue(self.settings["Spectral difference threshold"])
        self.spectral_difference_threshold_sb.setToolTip("Set the value the spectral difference threshold")
        self.spectral_difference_threshold_sb.valueChanged.connect(self.on_spectral_difference_threshold_sb_changed)
        # Add the number of iterations
        self.iterations_sb = QSpinBox(self)
        self.iterations_sb.setRange(1,900)
        self.iterations_sb.setValue(self.settings["Number of iterations"])
        self.iterations_sb.setToolTip("Set the number of iterations to be used to optimise the cross-correlation coefficient")
        self.iterations_sb.valueChanged.connect(self.on_iterations_sb_changed)
        # Independent y-axes
        self.independent_yaxes_cb = QCheckBox(self)
        self.independent_yaxes_cb.setToolTip("Check to use indpendent y-axes in the plot")
        self.independent_yaxes_cb.setText("")
        self.independent_yaxes_cb.setLayoutDirection(Qt.RightToLeft)
        if self.settings["Independent y-axes"]:
            self.independent_yaxes_cb.setCheckState(Qt.Checked)
        else:
            self.independent_yaxes_cb.setCheckState(Qt.Unchecked)
        self.independent_yaxes_cb.stateChanged.connect(self.on_independent_yaxes_cb_changed)
        # Optimise frequency scaling?
        self.optimise_frequency_scaling_cb = QCheckBox(self)
        self.optimise_frequency_scaling_cb.setToolTip("Use frequency scaling in optimisation")
        self.optimise_frequency_scaling_cb.setText("")
        self.optimise_frequency_scaling_cb.setLayoutDirection(Qt.RightToLeft)
        if self.settings["Optimise frequency scaling"]:
            self.optimise_frequency_scaling_cb.setCheckState(Qt.Checked)
        else:
            self.optimise_frequency_scaling_cb.setCheckState(Qt.Unchecked)
        self.optimise_frequency_scaling_cb.stateChanged.connect(self.on_optimise_frequency_scaling_cb_changed)
        #
        # Add a tab widget for the settings ######################################################################################
        #
        self.settingsTab = QTabWidget(self)
        self.settingsTab.addTab(self.spectrafile_le, "Experimental spectrum")
        self.settingsTab.addTab(self.scenario_cb, "Scenario")
        self.settingsTab.addTab(self.frequency_scaling_factor_sb, "Frequency scaling factor")
        self.settingsTab.addTab(self.iterations_sb, "No. of iterations")
        self.settingsTab.addTab(self.independent_yaxes_cb, "Independent y-axes")
        self.settingsTab.addTab(self.fitting_type_cb, "Fitting type")
        self.settingsTab.addTab(self.optimise_frequency_scaling_cb, "Optimise scaling")
        self.settingsTab.addTab(self.spectral_difference_threshold_sb, "Spectral difference threshold")
        self.settingsTab.addTab(self.baseline_cb, "Baseline removal?")
        self.settingsTab.addTab(self.hpfilter_lambda_sb, "HP Filter Lambda")
        label = QLabel("Options:", self)
        form.addRow(label,self.settingsTab)
        # END OF THE SETTINGS TAB #################################################################################################
        # Add Lorentzian widths table
        # get initial sigmas from the settings tab
        self.sigmas_cm1 = self.notebook.settingsTab.sigmas_cm1
        self.modes_selected = self.notebook.settingsTab.modes_selected
        self.frequencies_cm1 = self.notebook.settingsTab.frequencies_cm1
        self.intensities = self.notebook.settingsTab.intensities
        if len(self.modes_fitted) == 0 and len(self.modes_selected) > 0:
            self.modes_selected = [ False  for _ in self.modes_selected ]
        self.sigmas_tw = FixedQTableWidget(parent=self,rows=7)
        self.sigmas_tw.setToolTip("Choose the sigmas which will be used in the fitting")
        self.sigmas_tw.itemChanged.connect(self.on_sigmas_tw_itemChanged)
        self.sigmas_tw.setRowCount(max(8,len(self.sigmas_cm1)))
        self.sigmas_tw.setColumnCount(3)
        self.sigmas_tw.setHorizontalHeaderLabels(["   Sigma   \n(cm-1)", " Frequency \n(cm-1)", "  Intensity  \n(Debye2/Å2/amu)"])
        self.redraw_sigmas_tw()
        label = QLabel("Lorentzian widths:")
        label.setToolTip("The Lorentzian widths can be edited here.  If checked they will also be used in the optimisation of the cross-correlation between the experiment and calculated spectra")
        form.addRow(label,self.sigmas_tw)
        # Add a replot and recalculate button
        hbox = QHBoxLayout()
        self.replotButton1 = QPushButton("Replot")
        self.replotButton1.setToolTip("Recalculate the spectrum with the new sigma values")
        self.replotButton1.clicked.connect(self.replotButton1Clicked)
        hbox.addWidget(self.replotButton1)
        self.replotButton2 = QPushButton("Replot with frequency shift")
        self.replotButton2.setToolTip("Recalculate the spectrum with the new sigma values, including a shft in the frequencies to maximise the cross-correlation")
        self.replotButton2.clicked.connect(self.replotButton2Clicked)
        hbox.addWidget(self.replotButton2)
        # Add a fitting button
        self.fittingButton = QPushButton("Perform fitting")
        self.fittingButton.setToolTip("Attempt to fit the calculated spectrum to the experimental one")
        self.fittingButton.clicked.connect(self.fittingButtonClicked)
        hbox.addWidget(self.fittingButton)
        form.addRow(hbox)
        # Add a progress bar
        label = QLabel("Calculation progress", self)
        label.setToolTip("Show the progress of any calculations")
        self.progressbar = QProgressBar(self)
        form.addRow(label,self.progressbar)
        self.notebook.progressbars_add(self.progressbar)
        # Add output of the cross correlation coefficient
        hbox = QHBoxLayout()
        self.cross_correlation_le = QLineEdit(self)
        self.cross_correlation_le.setEnabled(False)
        self.cross_correlation_le.setText(f"{0.0}")
        self.lag_frequency_le = QLineEdit(self)
        self.lag_frequency_le.setEnabled(False)
        self.lag_frequency_le.setText(f"{0.0}")
        self.frequency_scaling_le = QLineEdit(self)
        self.frequency_scaling_le.setEnabled(False)
        self.frequency_scaling_le.setText("{}".format(self.settings["Frequency scaling factor"]))
        self.rmse_le = QLineEdit(self)
        self.rmse_le.setEnabled(False)
        self.rmse_le.setText(f"{0.0}")
        hbox.addWidget(self.cross_correlation_le)
        hbox.addWidget(self.lag_frequency_le)
        hbox.addWidget(self.frequency_scaling_le)
        hbox.addWidget(self.rmse_le)
        label = QLabel("X-correlation, shift/lag, frequency scale and rmse")
        label.setToolTip("The highest cross-correlation value and its associated frequency shift is shown followed by the spectral error")
        form.addRow(label, hbox)
        # Add the matplotlib figure to the bottom
        self.figure = matplotlib.figure.Figure()
        self.canvas = FigureCanvas(self.figure)
        self.canvas.setSizePolicy(QSizePolicy(QSizePolicy.Expanding,QSizePolicy.Expanding))
        self.toolbar = NavigationToolbar(self.canvas, self)
        self.toolbar.setSizePolicy(QSizePolicy(QSizePolicy.Fixed,QSizePolicy.Fixed))
        form.addRow(self.canvas)
        form.addRow(self.toolbar)
        vbox.addLayout(form)
        # finalise the layout
        self.setLayout(vbox)
        self.refreshRequired = True
        debugger.print("Finished:: Initialising")

    def on_iterations_sb_changed(self):
        """Handle changes in the iterations spin box value.

        This method is called when the iterations spin box value is changed. It updates the 'Number of iterations' setting with the new value and flags that a refresh is required.

        Parameters
        ----------
        None

        Returns
        -------
        None

        """        
        debugger.print("on_iterations_sb_changed")
        self.settings["Number of iterations"] = self.iterations_sb.value()
        self.refreshRequired = True
        return

    def on_independent_yaxes_cb_changed(self,value):
        """Handle the change of state for the independent Y-axes checkbox.

        Parameters
        ----------
        value : bool
            The new state of the checkbox. The actual value might not be used in the function but represents the change.

        Returns
        -------
        None

        Notes
        -----
        This function is triggered upon the change of the 'independent_yaxes_cb' checkbox state. It updates the settings dictionary with the new state of the 'Independent y-axes' option and sets the 'refreshRequired' flag to True, indicating that a refresh is needed to reflect the changes in the UI or data visualization.

        """        
        debugger.print("independent_yaxes_cb_changed",value)
        self.settings["Independent y-axes"] = self.independent_yaxes_cb.isChecked()
        self.refreshRequired = True
        return

    def on_optimise_frequency_scaling_cb_changed(self,value):
        """Handle changes in the frequency scaling optimization option.

        This method updates the 'Optimise frequency scaling' setting based on the state of a checkbox and marks that a refresh is required based on the new setting.

        Parameters
        ----------
        value : bool
            The value or state that triggered the change. This parameter is logged but not directly used to set the 'Optimise frequency scaling' option.

        Returns
        -------
        None

        """        
        debugger.print("optimise_frequency_scaling_cb_changed",value)
        self.settings["Optimise frequency scaling"] = self.optimise_frequency_scaling_cb.isChecked()
        self.refreshRequired = True
        return

    def on_spectrum_scaling_cb_changed(self,value):
        """Handle the change event of the spectrum scaling checkbox.

        This method is triggered whenever the spectrum scaling checkbox's state changes. It updates the settings to reflect the current state of the checkbox and flags that a refresh is required.

        Parameters
        ----------
        value : bool
            The new value of the spectrum scaling checkbox. The specific type and significance can vary depending on the implementation details of the checkbox and its handlers.

        Returns
        -------
        None

        """        
        debugger.print("on_spectrum_scaling_cb_changed",value)
        self.settings["Spectrum scaling"] = self.spectrum_scaling_cb.isChecked()
        self.refreshRequired = True
        return

    def on_hpfilter_lambda_sb_changed(self,value):
        """Handle changes to the high-pass filter lambda scrollbar value.

        Parameters
        ----------
        value : str
            The new value from the high-pass filter lambda scrollbar, intended to be converted to float.

        Returns
        -------
        None

        """        
        debugger.print("hpfilter_lambda_sb_changed",value)
        self.refreshRequired = True
        try:
            self.settings["HPFilter lambda"] = float(value)
        except Exception:
            print("Failed to convert to float", value)
        return

    def on_spectrum_scaling_factor_sb_changed(self,value):
        """Handle the change in spectrum scaling factor.

        Parameters
        ----------
        value : str
            The new value for the spectrum scaling factor. Expected to be convertible to float.

        Returns
        -------
        None

        Notes
        -----
        This function attempts to update the 'Spectrum scaling factor' setting with the new value after converting it to float. If the conversion fails, a message is printed indicating the failure.

        """        
        debugger.print("on_spectrum_scaling_factor_cb_changed",value)
        self.refreshRequired = True
        try:
            self.settings["Spectrum scaling factor"] = float(value)
        except Exception:
            print("Failed to convert to float", value)
        return

    def on_baseline_cb_changed(self,value):
        """Handle the change event of the baseline checkbox.

        This method is called when the baseline checkbox state changes. It updates the setting for baseline removal and marks refresh as required.

        Parameters
        ----------
        value : str
            The new value or state of the baseline checkbox. The actual data type and use of this parameter depend on the specific implementation and could vary.

        Returns
        -------
        None

        """        
        debugger.print("on_baseline_cb_changed",value)
        self.refreshRequired = True
        self.settings["Baseline removal"] = self.baseline_cb.isChecked()
        return

    def on_spectral_difference_threshold_sb_changed(self,value):
        """Handle changes in spectral difference threshold setting.

        This function is called when the 'Spectral difference threshold' setting is changed.
        It updates the setting value, converting it to a float, and sets a flag indicating that a refresh is required.

        Parameters
        ----------
        value : str
            The new value for the 'Spectral difference threshold' setting. The function attempts to convert this to a float.

        Returns
        -------
        None

        """        
        debugger.print("on_spectral_difference_threshold_sb_changed",value)
        self.refreshRequired = True
        try:
            self.settings["Spectral difference threshold"] = float(value)
        except Exception:
            print("Failed to convert to float", value)
        return

    def on_frequency_scaling_factor_sb_changed(self,value):
        """Handle changes in the frequency scaling factor setting.

        This function changes the frequency scaling factor setting. It updates the setting with the new value and marks a refresh as necessary. It also handles conversion errors gracefully.

        Parameters
        ----------
        value : str
            The new value for the frequency scaling factor setting. The function attempts to convert this to a float.

        Returns
        -------
        None

        Notes
        -----
        - The `self.refreshRequired` flag is set to True regardless of whether the conversion succeeds.

        """        
        debugger.print("on_frequency_scaling_factor_cb_changed",value)
        self.refreshRequired = True
        try:
            self.settings["Frequency scaling factor"] = float(value)
        except Exception:
            print("Failed to convert to float", value)
        return

    def replotButton1Clicked(self):
        """Handle the click event on the replot button.

        This method is triggered when the replot button is clicked. It sets up the necessary flags for the replot operation and invokes the refresh method to update the plot.

        Parameters
        ----------
        None

        Returns
        -------
        None

        Notes
        -----
        This method modifies the instance attributes to reflect that a replot is required without a frequency shift, and it bookmarks the last button pressed as 'replotButton1Clicked'. It also calls the `self.refresh` method to initiate the replotting process.

        """        
        debugger.print("Start:: replotButton1Clicked")
        self.refreshRequired = True
        self.plot_frequency_shift = False
        self.lastButtonPressed = self.replotButton1Clicked
        self.refresh()
        debugger.print("Finished:: replotButton1Clicked")
        return

    def replotButton2Clicked(self):
        """Handle the event when the second replot button is clicked.

        This method updates various attributes to indicate that a refresh of the plot is required,
        specifically with a frequency shift. It records that this button was the last one pressed
        and then initiates the refresh process.

        Parameters
        ----------
        None

        Returns
        -------
        None

        """        
        debugger.print("Start:: replotButton2Clicked")
        self.refreshRequired = True
        self.plot_frequency_shift = True
        self.lastButtonPressed = self.replotButton2Clicked
        self.refresh()
        debugger.print("Finished:: replotButton2Clicked")
        return

    def plot(self,experiment,xs,ys,legends,label):
        """Plot experimental and calculated data on the same or separate y-axes.

        This method plots provided experimental and calculated data sets using Matplotlib. It supports plotting
        data with independent y-axes and applying frequency shift and scaling to the x-axis. The plot appearance,
        including color mapping, line width, and labels, is customizable.

        Parameters
        ----------
        experiment : array_like
            The experimental data to be plotted. If empty, no experimental data is plotted.
        xs : list of array_like
            A list containing the x-axis data for each data set.
        ys : list of array_like
            A list containing the y-axis data for each data set to be plotted alongside the x data.
        legends : list of str
            A list containing the labels for each data set.
        label : str
            The label for the plot, typically used as a part of y-axis label.

        Returns
        -------
        None

        Raises
        ------
        ValueError
            If `xs`, `ys`, and `legends` lists do not have the same length.

        """        
        # Plot the experimental values on the left y-axis
        # Plot all the others in xs, ys on the right x-axis
        debugger.print("Start:: plot")
        self.subplot1 = None
        self.figure.clf()
        import matplotlib.pyplot as plt
        cmap = plt.get_cmap("tab10")
        if self.frequency_units == "wavenumber":
            xlabel = r"Frequency $\mathdefault{(cm^{-1})}$"
            scale = 1.0
        else:
            xlabel = r"THz"
            scale = 0.02998
        self.subplot1 = self.figure.add_subplot(111)
        if self.settings["Independent y-axes"]:
            self.subplot2 = self.subplot1.twinx()
        else:
            self.subplot2 = self.subplot1
        cmap_index = 0
        lines = []
        if self.plot_frequency_shift:
            lag = float(self.lag_frequency_le.text())
            scale_calc = scale * self.settings["Frequency scaling factor"]
        else:
            lag = 0.0
            # scale_calc = scale
            scale_calc = scale * self.settings["Frequency scaling factor"]
        for x,y,legend in zip(xs,ys,legends):
            if y is not None:
               x = np.array(x)
               line, = self.subplot1.plot(lag+scale_calc*x,y,lw=2, color=cmap(cmap_index), label=legend )
               lines.append(line)
               cmap_index += 1
        if len(experiment) > 0:
            # Use the x variables from the previous xs, ys
            line, = self.subplot2.plot(scale*x,experiment,lw=2, color=cmap(cmap_index), label="Experiment" )
            lines.append(line)
        labels = [l.get_label() for l in lines]
        if self.settings["Independent y-axes"]:
            self.subplot2.set_ylabel("Experiment")
            self.subplot2.set_ylim(bottom=0.0)
            self.subplot1.set_ylabel("Calculated "+self.plot_label )
        else:
            self.subplot1.set_ylabel(self.plot_label )
        self.subplot1.set_xlabel(xlabel)
        self.subplot1.set_ylim(bottom=0.0)
        self.subplot1.legend(lines, labels, loc="best")
        self.subplot1.set_title(self.settings["Plot title"])
        self.canvas.draw_idle()
        debugger.print("Finished:: plot")

    def fittingButtonClicked(self):
        """Handle the click event on the fitting button.

        This method toggles the fitting process on and off. When clicked,
        it changes the button's text accordingly to reflect the current
        state - either "Perform fitting" or "Interrupt fitting". It also
        triggers re-plotting and refreshes the UI. If the fitting process
        starts, it performs optimization and then resets the button's state
        to allow a new fitting process.

        Parameters
        ----------
        None

        Returns
        -------
        None

        """        
        debugger.print("Start:: fittingButtonClicked")
        self.refreshRequired = True
        if self.calculationInProgress:
            self.fittingButton.setText("Perform fitting")
            self.calculationInProgress = False
        else:
            self.fittingButton.setText("Interupt fitting")
            self.calculationInProgress = True
        debugger.print("replotButton2Clicked",self.refreshRequired)
        self.refresh()
        self.replot()
        self.optimiseFit()
        self.fittingButton.setText("Perform fitting")
        self.calculationInProgress = False
        debugger.print("Finished:: fittingButtonClicked")
        return

    def optimiseFit(self):
        # Optimise the fit of the first scenario to the experimental data
        # First determine who many variables we have
        """Optimise the fit based on the initial and adjusted sigmas and potential frequency scaling.

        Instructions for determining the best fitting parameters for the model are conducted
        within the constraints of predefined modes. The function iterates to minimize the discrepancy
        between modeled and experimental data. It employs the Nelder-Mead optimization method for
        solving the problem.

        Parameters
        ----------
        None

        Returns
        -------
        list or OptimizeResult
            The final optimization point as a list if no sigma has been selected for optimisation,
            or as an `OptimizeResult` object returned by `scipy.optimize.minimize`. The return type
            depends on whether the optimization process is executed.

        Notes
        -----
        The function uses several attributes of the class instance it belongs to:
        - `modes_fitted` to determine which modes are fitted.
        - `sigmas_cm1` as initial sigma points for optimization.
        - `settings` to check if frequency scaling should be optimised and to determine the number of iterations for the optimization.
        - `optimiseFunction` as the function to be minimized.

        It modifies the `functionCalls` attribute by setting it to 0 at the beginning and uses
        `fit_list` to store the indexes of modes that are being fitted.

        """        
        debugger.print("Start:: optimiseFit")
        self.functionCalls = 0
        self.fit_list = []
        for mode,fitted in enumerate(self.modes_fitted):
            if fitted:
                self.fit_list.append(mode)
        initial_point = [ self.sigmas_cm1[i] for i in self.fit_list ]
        # Append a scaling option
        if self.settings["Optimise frequency scaling"]:
            initial_point.append(self.settings["Frequency scaling factor"])
        nvariables = len(initial_point)
        if nvariables > 0:
            final_point = minimize(self.optimiseFunction, initial_point, method="nelder-mead", options={"xatol":0.01, "disp":True, "maxfev":nvariables+nvariables*self.settings["Number of iterations"]} )
        else: 
            print("No sigmas have been selected for optimisation")
            final_point = []
        debugger.print("Finished:: optimiseFit")
        return final_point

    def optimiseFunction(self,variables) :
        """Optimise the fitting function based on given variables.

        This method adjusts internal settings based on an array of variables to
        optimise the fitting function according to the predefined criteria
        (e.g., minimize cross-correlation or spectral difference). It updates
        GUI components and recalculates the function value based on the new
        configuration.

        Parameters
        ----------
        variables : list or array-like
            The variables used for optimization. This could include sigma values
            for each element to be fitted and optionally a scaling factor for
            frequency scaling, depending on the configuration.

        Returns
        -------
        float
            The function value after optimization, which is either the negative
            cross-correlation or the root mean square error (RMSE), based on the
            fitting type setting.

        Notes
        -----
        - `calculationInProgress` is a boolean attribute of the
          object that indicates whether a calculation can proceed.
        - Manipulates GUI components such as `fittingButton` and elements within
          `notebook.settingsTab`, thus requires the GUI to be in a responsive state.
        - Uses global `debugger` object for logging progress and results.
        - The method updates the internal state of the object, such as the
          `functionCalls` counter and sigma values for elements being fitted.

        Raises
        ------
        Exception
            If the calculation is not in progress (indicated by
            `calculationInProgress` being `False`), it immediately returns a
            predefined extremely negative value (e.g., -9.9e+99).

        """        
        if not self.calculationInProgress:
            return -9.9E99
        debugger.print("Start:: optimiseFunction",variables)
        self.functionCalls += 1
        nvariables = len(variables)
        self.fittingButton.setText("Interrupt fitting ({}/{})".format(self.functionCalls,nvariables+1+nvariables*self.settings["Number of iterations"]))
        if self.settings["Optimise frequency scaling"]:
            sigmas = variables[:-1]
            scaling_factor = variables[-1]
            self.settings["Frequency scaling factor"] = scaling_factor
        else:
            sigmas = variables
            scaling_factor = self.settings["Frequency scaling factor"]
        for index,sigma in zip(self.fit_list,sigmas):
            self.sigmas_cm1[index] = sigma
            self.redraw_sigmas_tw()
            self.notebook.settingsTab.sigmas_cm1[index] = sigma
            self.notebook.settingsTab.redraw_output_tw()
            self.notebook.settingsTab.requestRefresh()
        self.refresh(force=True)
        self.replot()
        # Returning the best correlation but made negative because we need to minimise
        if self.settings["Fitting type"] == "Minimise x-correlation":
            function_value = -1.0*self.xcorr0
        elif self.settings["Fitting type"] == "Minimise spectral difference":
            function_value = self.rmse
        debugger.print("optimiseFunction - xcorr0,rmse",self.xcorr0,self.rmse)
        debugger.print("Finished:: optimiseFunction",function_value)
        return function_value


    def redraw_sigmas_tw(self):
        """Redraw the table widget for sigmas.

        This method updates the table widget `sigmas_tw` used for displaying sigma, frequency, and intensity
        information. It checks if the data lists (`sigmas_cm1`, `frequencies_cm1`, and `intensities`) are non-empty
        and then proceeds to populate the table. Each row in the table corresponds to a unique sigma, frequency,
        and intensity value. The method also handles item selection and editing properties based on the state of
        `modes_selected` and `modes_fitted` attributes.

        Parameters
        ----------
        None

        Returns
        -------
        None

        Notes
        -----
        - `sigmas_cm1`, `frequencies_cm1`, and `intensities` are lists that store sigma, frequency, and intensity data, respectively.
        - `modes_selected` and `modes_fitted` are lists that store boolean values indicating whether a mode is selected and
          whether it is fitted, respectively.

        """        
        debugger.print("Start:: redraw_sigmas_tw")
        if len(self.sigmas_cm1) <= 0:
            debugger.print("Finished:: redraw_sigmas_tw")
            return
        self.sigmas_tw.blockSignals(True)
        self.sigmas_tw.setRowCount(len(self.sigmas_cm1))
        for i,(f,sigma,intensity) in enumerate(zip(self.frequencies_cm1, self.sigmas_cm1, self.intensities)):
            # Sigma and check / unchecked column
            items = []
            itemFlags = []
            item = QTableWidgetItem(f"{sigma:.2f}")
            if self.modes_selected[i]:
                if self.modes_fitted[i]:
                    item.setCheckState(Qt.Checked)
                else:
                    item.setCheckState(Qt.Unchecked)
                itemFlags.append( item.flags() & Qt.NoItemFlags | Qt.ItemIsUserCheckable | Qt.ItemIsEnabled | Qt.ItemIsSelectable | Qt.ItemIsEditable )
                otherFlags = item.flags() & Qt.NoItemFlags | Qt.ItemIsEnabled
            else:
                #itemFlags.append( item.flags() | Qt.ItemIsUserCheckable | Qt.ItemIsEnabled & ~Qt.ItemIsSelectable & ~Qt.ItemIsEditable )
                itemFlags.append( item.flags() & Qt.NoItemFlags | Qt.ItemIsUserCheckable | Qt.ItemIsEnabled )
                item.setCheckState(Qt.Unchecked)
                otherFlags = item.flags() & Qt.NoItemFlags
            items.append(item)
            # Frequency column cm-1
            items.append(QTableWidgetItem(f"{f:.4f}" ) )
            itemFlags.append( otherFlags )
            # Intensity column Debye2/Angs2/amu
            items.append(QTableWidgetItem(f"{intensity:.4f}" ) )
            itemFlags.append( otherFlags )
            for j,(item,flag) in enumerate(zip(items,itemFlags)):
                item.setFlags(flag)
                item.setTextAlignment(int(Qt.AlignHCenter | Qt.AlignVCenter))
                self.sigmas_tw.setItem(i, j, item )
        self.sigmas_tw.resizeColumnsToContents()
        # Release the block on signals for the frequency output table
        self.sigmas_tw.blockSignals(False)
        QCoreApplication.processEvents()
        debugger.print("Finished:: redraw_sigmas_tw")

    def on_fitting_type_cb_activated(self,index):
        # Change in fitting type
        """Handle the activation of a fitting type option.

        This method updates the current settings to reflect the selected fitting type based on the provided index. It also marks the state as requiring a refresh.

        Parameters
        ----------
        index : int
            The index of the selected fitting type in the fitting type definitions list.

        Returns
        -------
        None

        """        
        debugger.print("on_fitting_type_cb_activated", index)
        self.refreshRequired = True
        self.settings["Fitting type"] = self.fitting_type_definitions[index]

    def on_scenario_cb_activated(self,index):
        # Change in Scenario to be used for fitting
        """Handle scenario combobox activation.

        This function is triggered when a scenario combobox option is selected. It logs the activation with the selected index, marks that a refresh is required, and updates the settings to reflect the new scenario.

        Parameters
        ----------
        index : int
            The index of the activated scenario.

        Returns
        -------
        None

        """        
        debugger.print("on_scenario_cb_activated", index)
        self.refreshRequired = True
        self.settings["Scenario index"] = index

    def on_spectrafile_le_return(self):
        # Handle a return in the experimental file name line editor
        """Handle file selection for experimental spectra.

        This method is executed when an action associated with the `spectrafile_le` widget is performed. It prompts the user to select a file if the currently specified file does not exist or cannot be found. It updates the application's settings with the new file name, if a file is selected.

        Parameters
        ----------
        None

        Returns
        -------
        None

        Notes
        -----
        The method first checks if the file specified in the `spectrafile_le` widget exists. If it does not, a file dialog is opened for the user to select a file. If a file is selected, the method updates the application's settings and the text of the `spectrafile_le` widget to reflect the new file name. Additionally, it sets a flag indicating that a refresh is required and that the experimental file needs to be read again. Finally, it triggers the last button pressed action.

        This method utilizes `os.path.isfile` to check file existence, `QFileDialog` for file selection. 

        """        
        debugger.print("Start:: on_spectrafile_le_return")
        file_name = self.spectrafile_le.text()
        if not os.path.isfile(file_name):
            qfd = QFileDialog(self)
            qfd.setDirectory(self.notebook.mainTab.directory)
            file_name, _ = qfd.getOpenFileName(self,"Open the experimental file","","Excel (*.xlsx);; Csv (*.csv);; All Files (*)")
        if not os.path.isfile(file_name):
            debugger.print("Finished:: on_spectrafile_le_return")
            return
        self.settings["Experimental file name"] = file_name
        self.spectrafile_le.setText(self.settings["Experimental file name"])
        debugger.print("new file name", self.settings["Experimental file name"])
        self.refreshRequired = True
        self.experimental_file_has_been_read = False
        # redo the plot if a return is pressed
        self.lastButtonPressed()
        debugger.print("Finished:: on_spectrafile_le_return")
        return

    def on_spectrafile_le_changed(self,text):
        """Handle changes to the spectra file input field.

        This method is triggered when the text of the spectrum file input field changes. It updates the 'Experimental file name' setting with the new text, marks the current experimental file as unread, and indicates that a refresh is required.

        Parameters
        ----------
        text : str
            The new text from the spectra file input field. Not directly used after method update.

        Returns
        -------
        None

        Notes
        -----
        The actual text used in processing is fetched directly from the `spectrafile_le` attribute, not from the `text` argument.

        """        
        debugger.print("on_spectrafile_le_changed", text)
        text = self.spectrafile_le.text()
        self.refreshRequired = True
        self.settings["Experimental file name"] = text
        self.experimental_file_has_been_read = False

    def on_sigmas_tw_itemChanged(self, item):
        """Handle item changed event in a table widget for sigma values.

        This method updates mode fitting status and sigma values based on changes in a table widget,
        redraws the table widget, and refreshes output based on the modifications. It also handles error
        for invalid numeric input for sigma values. 

        Parameters
        ----------
        item : QTableWidgetItem
            The table widget item that has been changed.

        Returns
        -------
        None

        Side Effects
        ------------
        - Updates internal data structures storing mode fitting statuses and sigma values.
        - Triggers UI update methods to reflect changes in the table widget and other dependent UI components.
        - Blocks and unblocks signals from the table widget to prevent recursive item change events during updates.
        - Processes all pending events for the application to ensure UI updates are immediately applied.

        """        
        debugger.print("Start:: on_sigmas_tw_itemChanged", item)
        self.sigmas_tw.blockSignals(True)
        debugger.print("on_sigmas_tw_itemChanged)", item.row(), item.column() )
        col = item.column()
        row = item.row()
        if col == 0:
            # If this is the first column alter the check status but reset the sigma value
            if item.checkState() == Qt.Checked:
                self.modes_fitted[row] = True
            else:
                 self.modes_fitted[row] = False
            try:
                new_value = float(item.text())
                self.sigmas_cm1[row] = new_value
                self.redraw_sigmas_tw()
                self.notebook.settingsTab.sigmas_cm1[row] = new_value
                self.notebook.settingsTab.requestRefresh()
                self.notebook.settingsTab.redraw_output_tw()
            except Exception:
                 print("Failed to convert to float",item.txt())
        elif col == 1:
            self.redraw_sigmas_tw()
        else:
            self.redraw_sigmas_tw()
        self.refreshRequired = True
        QCoreApplication.processEvents()
        debugger.print("Finished:: on_sigmas_tw_itemChanged", item)
        return


    def print_settings(self):
        """Print the settings attributes of the instance.

        Parameters
        ----------
        None

        Returns
        -------
        None

        Prints the key and value of each setting in the instance's `settings` attribute, with the help of the `debugger` print method.

        """        
        debugger.print("print_settings")
        for key in self.settings:
            debugger.print(key, self.settings[key])
        return

    def replot(self):
        """Replots the experimental spectrum along with calculated frequencies and spectra.

        This method will initially check if there is a resampled experimental spectrum. If such a spectrum exists, the spectrum is resampled. Subsequently, it calls the plot method with the updated data along with other related information such as legends and labels.

        Parameters
        ----------
        None

        Returns
        -------
        None

        """        
        debugger.print("Start:: replot")
        self.xaxis = np.array(self.calculated_frequencies[0])
        if len(self.resampled_experimental_spectrum) > 0:
            self.resampled_experimental_spectrum = resample_experimental_spectrum(self.xaxis,
                                                       self.experimental_frequencies,
                                                       self.experimental_spectrum,
                                                       baseline_removal = self.settings["Baseline removal"],
                                                       HPFilter_lambda  = self.settings["HPFilter lambda"],
                                                       debug=debugger.state())
        self.plot(self.resampled_experimental_spectrum,self.calculated_frequencies,self.calculated_spectra,self.scenario_legends,self.plot_label)
        debugger.print("Finished:: replot")
        return

    def refresh(self,force=False):
        """Refresh the state of the application based on user-defined conditions and settings.

        Parameters
        ----------
        force : bool, optional
            If True, the refresh proceeds regardless of whether it is deemed necessary based on the internal state. Defaults to False.

        Returns
        -------
        None

        Notes
        -----
        The refresh process involves several key steps including checking if a refresh is required, updating plotting data from settings, reading an experimental file, setting GUI components based on notebook scenarios, and recalculating spectra. If the refresh is aborted either due to being unnecessary or because there are no frequencies, it returns early. Additionally, if `force` is True, it ignores the check for whether a refresh is required and proceeds with the refresh.

        After updating relevant class attributes and GUI components, the function recalculates cross-correlation, root-mean-square error (RMSE) between experimental and calculated spectra, and updates the display with these metrics.

        """        
        debugger.print("Start:: refresh", force)
        if not self.refreshRequired and not force:
            debugger.print("refresh aborted", self.refreshRequired,force)
            debugger.print("Finished:: refresh", force)
            return
        self.frequencies_cm1 = self.notebook.settingsTab.frequencies_cm1
        if np.sum(self.frequencies_cm1) < 1.0E-10:
            debugger.print("refresh aborted there are no frequencies")
            debugger.print("Finished:: refresh", force)
            return
        #
        # Flag all the scenarios as needing an update
        #
        for scenario in self.notebook.scenarios:
            scenario.requestRefresh()
        self.notebook.settingsTab.requestRefresh()
        #
        # Now refresh the plotting tab 
        #
        self.notebook.plottingTab.refresh()
        #
        # Block signals during refresh
        # 
        #for w in self.findChildren(QWidget):
        #    w.blockSignals(True)
        # 
        # use the settings values to initialise the widgets
        #
        self.spectrafile_le.setText(self.settings["Experimental file name"])
        #
        # Only read the file again if we have to
        #
        if not self.experimental_file_has_been_read :
            self.experimental_frequencies,self.experimental_spectrum = read_experimental_file(self.settings["Experimental file name"],debug=debugger.state())
        if len(self.experimental_frequencies) > 0:
            self.experimental_file_has_been_read = True
        self.plot_label = self.notebook.plottingTab.settings["Plot type"]
        self.scenario_legends = [ scenario.settings["Legend"] for scenario in self.notebook.scenarios ]
        self.scenario_cb.clear()
        self.scenario_cb.addItems(self.scenario_legends)
        self.scenario_cb.setCurrentIndex(self.settings["Scenario index"])
        vs_cm1 = self.notebook.plottingTab.vs_cm1
        self.calculated_spectra = [ scenario.get_result(vs_cm1, self.plot_label) for scenario in self.notebook.scenarios ]
        self.calculated_spectrum = self.calculated_spectra[self.settings["Scenario index"]]
        debugger.print("refresh scenario index" , self.settings["Scenario index"])
        self.iterations_sb.setValue(self.settings["Number of iterations"])
        self.frequency_scaling_factor_sb.setValue(self.settings["Frequency scaling factor"])
        if self.settings["Independent y-axes"]:
            self.independent_yaxes_cb.setCheckState(Qt.Checked)
        else:
            self.independent_yaxes_cb.setCheckState(Qt.Unchecked)
        # 
        # If the sigmas are not set return
        #
        self.sigmas_cm1 = self.notebook.settingsTab.sigmas_cm1
        if len(self.sigmas_cm1) < 1:
            return
        self.modes_selected = self.notebook.settingsTab.modes_selected
        self.intensities = self.notebook.settingsTab.intensities
        if len(self.modes_fitted) == 0 and len(self.modes_selected) > 0:
            self.modes_fitted = [ False  for _ in self.modes_selected ]
        self.redraw_sigmas_tw()
        # Resample the spectrum
        self.calculated_frequencies = [ vs_cm1 for scenario in self.notebook.scenarios ]
        self.xaxis = self.calculated_frequencies[0]
        self.rmse = 0.0
        self.lag  = 0.0
        self.xcorr0  = 1.0
        self.xcorr1  = 1.0
        if len(self.experimental_spectrum) > 0:
            self.resampled_experimental_spectrum = resample_experimental_spectrum(self.xaxis,
                                                       self.experimental_frequencies,
                                                       self.experimental_spectrum,
                                                       baseline_removal = self.settings["Baseline removal"],
                                                       HPFilter_lambda  = self.settings["HPFilter lambda"],
                                                       debug=debugger.state())
            scaling_factor = self.settings["Frequency scaling factor"]
            spectral_threshold=self.settings["Spectral difference threshold"],
            self.lag,self.xcorr0,self.xcorr1 = calculateCrossCorrelation(self.xaxis,
                                                                     self.calculated_spectrum,
                                                                     self.resampled_experimental_spectrum,
                                                                     scaling_factor=scaling_factor,
                                                                     debug=debugger.state())
            self.rmse = calculateSpectralDifference(self.xaxis,
                                                self.calculated_spectrum,
                                                self.resampled_experimental_spectrum,
                                                scaling_factor=scaling_factor,
                                                spectral_threshold=spectral_threshold,
                                                debug=debugger.state())
        self.cross_correlation_le.setText(f"{self.xcorr0:6.4f}")
        self.lag_frequency_le.setText(f"{self.lag:8.2f}")
        self.frequency_scaling_le.setText("{:8.2f}".format(self.settings["Frequency scaling factor"]))
        self.rmse_le.setText(f"{self.rmse:.2e}")
        self.replot()
        #
        # Unblock signals after refresh
        # 
        for w in self.findChildren(QWidget):
            w.blockSignals(False)
        self.refreshRequired = False
        debugger.print("Finished:: refresh", force)
        return

    def requestRefresh(self):
        """Request a refresh of the interface.

        Parameters
        ----------
        None

        Returns
        -------
        None

        """
        debugger.print("Start:: requestRefresh")
        self.refreshRequired = True
        debugger.print("Finished:: requestRefresh")

