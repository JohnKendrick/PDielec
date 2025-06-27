#
# Copyright 2024 John Kendrick & Andrew Burnett
#
# This file is part of PDielec
#
# This program is free software; you can redistribute it and/or modify
# it under the terms of the MIT License
#
# This program is distributed in the hope that it will be useful,
# but WITHOUT ANY WARRANTY; without even the implied warranty of
# MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.
#
# You should have received a copy of the MIT License along with this program, if not see https://opensource.org/licenses/MIT
#
"""Materials DataBase.

An interface to the spreadsheet which holds materials data
"""

import math
import os
import sys

import numpy as np
import openpyxl as xl

from PDielec import Calculator, DielectricFunction
from PDielec import __file__ as PDielec_init_filename
from PDielec.UnitCell import UnitCell
from PDielec.Utilities import Debug


class MaterialsDataBase:
    """A class for managing a database of materials' properties.

    This database is initialized from an Excel spreadsheet which contains various material properties including names, densities, refractive indices, and permittivities, among others.
    Additional functionalities include validation checks, retrieval of sheet names, material information, and specific data based on the property of interest (e.g., constant permittivity, tabulated refractive index).
    The getMaterial() method returns a material with a dielectric function of the appropriate type.
    There are routines which read (process) the data stored for the following dielectric functions:

    - constant refractive index
    - constant permittivity
    - tabulated refractive index (may be 1, 3 or 6 parameters for isotropic, uniaxial or anisotropic)
    - tabulated permittivity (may be 1, 3 or 6 parameters for isotropic, uniaxial or anisotropic)
    - Lorentz-Drude
    - FPSQ (Four parameters semi-quantum model)
    - Sellmeier
    
    Further information can be found in the following classes and their sub-classes:

    - :class:`~PDielec.Materials.Material`
    - :class:`~PDielec.DielectricFunction.DielectricFunction`

    Parameters
    ----------
    filename : str
        The filename of the spreadsheet/database.
    debug : bool, optional
        Set to true for additional debugging information

    Attributes
    ----------
    filename : str
        The path to the Excel spreadsheet containing materials data.
    sheetNames : list
        A list of strings representing the names of the sheets within the Excel spreadsheet.
    cache : dictionary
        A dictionary of materials that have been read from the database
    debug : bool, optional
        A flag indicating whether debugging information should be printed. Default is False.

    Methods
    -------
    __init__(filename, debug=False)
        Initializes the MaterialsDataBase class with a given Excel spreadsheet and a debug flag.
    getFileName()
        Returns the filename of the Excel spreadsheet being used as the database.
    valid()
        Checks if the spreadsheet is a valid materials database based on certain criteria.
    getSheetNames()
        Retrieves a sorted list of sheet names within the spreadsheet, including additional predefined names.
    getMaterial(sheet)
        Returns a material object based on the data in a given sheet of the Excel spreadsheet.
    readConstantRefractiveIndex(sheet, worksheet, density)
        Reads constant refractive index data for a given material from the spreadsheet.
    readConstantPermittivity(sheet, worksheet, density)
        Reads constant permittivity data for a given material from the spreadsheet.
    readTabulatedRefractiveIndex(sheet, worksheet, density)
        Reads tabulated refractive index data for a given material from the spreadsheet.
    readTabulatedSpectroscopy(sheet, worksheet, density)
        Reads tabulated n and alpha (in cm-1) data for a given material from the spreadsheet.
    readTabulatedPermittivity(sheet, worksheet, density)
        Reads tabulated permittivity data for a given material from the spreadsheet.
    readLorentzDrude(sheet, worksheet, density, unitCell)
        Reads Drude-Lorentz model parameters for a given material from the spreadsheet.
    readFPSQ(sheet, worksheet, density, unitCell)
        Reads FPSQ model parameters for a given material from the spreadsheet.
    readSellmeier(sheet, worksheet, density, unitCell)
        Reads Sellmeier model parameters for a given material from the spreadsheet.

    """

    def __init__(self,filename, debug=False):
        """Initialise a database of material properties using the excel spreadsheet filename.

        Parameters
        ----------
        filename : str
            The filename of the spreadsheet/database.
        debug : bool, optional
            Set to true for additional debugging information

        """
        global debugger
        debugger = Debug(debug,"MaterialsDataBase")
        debugger.print("Start:: initialise")
        if len(filename)> 5 and (filename.endswith("xlsx") or filename.endswith("XLSX")) and os.path.isfile(filename):
            self.filename = os.path.relpath(filename)
            self.workbook = xl.load_workbook(self.filename,data_only=True)
            self.sheetNames = self.workbook.sheetnames
            debugger.print("Sheet names:: ",self.sheetNames)
            # Close the work book while it is not in use
            # workbook.close()
        else:
            # Try opening the default database
            PDielec_Directory = os.path.dirname(PDielec_init_filename)
            filename  = os.path.join(PDielec_Directory, "MaterialsDataBase.xlsx")
            filename  = os.path.relpath(filename)
            if os.path.isfile(filename):
                self.filename = filename
                self.workbook = xl.load_workbook(self.filename,data_only=True)
                self.sheetNames = self.workbook.sheetnames
                debugger.print("Sheet names from default database ",self.sheetNames)
                # Close the work book while it is not in use
                # workbook.close()
            else:
                self.filename = None
                self.sheetNames = None
                print("  Error: MaterialsDataBase filename not valid",filename)
        self.cache = {}
        debugger.print("Finished:: initialise")
        return

    def getFileName(self):
        """Return the filename.

        Parameters
        ----------
        None

        Returns
        -------
        str
            The filename.

        """
        return self.filename

    def valid(self):
        """Test to see if the spreadsheet is a valid materials database.

        Parameters
        ----------
        None

        Returns
        -------
        bool
            True if the spreadshee is 'valid'

        Notes
        -----
        None

        """
        result = False
        if "Information" in self.sheetNames[0]:
            result = True
        return result

    def getSheetNames(self):
        """Return a list of the sheetnames in the database.

        As well as the sheets in the database, there are some default materials which will be added
        to the list of materials and which are provided internally by this module.

        Parameters
        ----------
        None

        Returns
        -------
        list
            A list of the sheet names present in the database.

        """
        # First take a copy of the sheetnames ignoring the first (Information)
        fullList = []
        if self.sheetNames is not None:
            fullList = self.sheetNames[1:].copy()
        # Append any in-built materials
        # This list is taken from the original powder code before version 8.0
        if "air" not in fullList:
            fullList.append("air")
        if "vacuum" not in fullList:
            fullList.append("vacuum")
        if "ptfe" not in fullList:
            fullList.append("ptfe")
        if "ldpe" not in fullList:
            fullList.append("ldpe")
        if "mdpe" not in fullList:
            fullList.append("mdpe")
        if "kbr" not in fullList:
            fullList.append("kbr")
        if "nujol" not in fullList:
            fullList.append("nujol")
        debugger.print("getSheetNames:: ",fullList)
        return sorted(fullList, key=lambda s: s.casefold())

    def getMaterial(self,sheet):
        """Return a material object based on the data in sheet (an excel sheet).

        If one of the following is requested: air, vacuum, ptfe, ldpe, mdpe, kbr, nujol, then
        the material is created even if it is not in the database.

        Parameters
        ----------
        sheet : Excel sheet
            The excel sheet containing the material data.

        Returns
        -------
        Material object
            The material object created from the excel sheet data.

        """
        debugger.print("getMaterial:: ",sheet)
        # Lets see if the material is in the cache
        if sheet in self.cache:
            debugger.print("getMaterial:: using the cache")
            return self.cache[sheet]
        # Define a set of back-up materials that the program can use even if the sheet name is not in the spreadsheet
        if self.sheetNames is None or sheet not in self.sheetNames:
            if sheet == "air":
                material = Constant("air",permittivity=1.0,density=0.001225)
            elif sheet == "vacuum":
                material = Constant("vacuum",permittivity=1.0,density=0.0)
            elif sheet == "ptfe":
                material = Constant("ptfe",permittivity=1.0,density=2.2)
            elif sheet == "ldpe":
                material = Constant("ldpe",permittivity=2.25,density=0.925)
            elif sheet == "mdpe":
                material = Constant("mdpe",permittivity=2.25,density=0.933)
            elif sheet == "kbr":
                material = Constant("kbr",permittivity=2.25,density=2.75)
            elif sheet == "nujol":
                material = Constant("nujol",permittivity=2.155,density=0.838)
            else:
                print("Error in getMaterial sheet ",sheet," not in self.sheetNames",self.sheetNames,file=sys.stderr)
                material = Constant("vacuum",permittivity=1.0,density=0.0)
            return material
        # Carry on with the spreadsheet
        # workbook = xl.load_workbook(self.filename,data_only=True)
        worksheet = self.workbook[sheet]
        unitCell = None
        avector = bvector = cvector = None
        a = b = c = alpha = beta = gamma = None
        for i in range(20):
            cell1 = "G"+str(i+1)
            cell2 = "H"+str(i+1)
            token = worksheet[cell1].value
            if token is not None:
                token = token.lower()
                if "entry" in token:
                    entry = worksheet[cell2].value.lower()
                elif "density" in token:
                    density = float(worksheet[cell2].value)
                elif "a_vector" in token:
                    avector = [ float(cell.value) for cell in [ worksheet["I"+str(i+1)], worksheet["J"+str(i+1)], worksheet["K"+str(i+1)] ] ]
                elif "b_vector" in token:
                    bvector = [ float(cell.value) for cell in [ worksheet["I"+str(i+1)], worksheet["J"+str(i+1)], worksheet["K"+str(i+1)] ] ]
                elif "c_vector" in token:
                    cvector = [ float(cell.value) for cell in [ worksheet["I"+str(i+1)], worksheet["J"+str(i+1)], worksheet["K"+str(i+1)] ] ]
                elif token == "a:":
                    a = float(worksheet[cell2].value)
                elif token == "b:":
                    b = float(worksheet[cell2].value)
                elif token == "c:":
                    c = float(worksheet[cell2].value)
                elif "alpha" in token:
                    alpha = float(worksheet[cell2].value)
                elif "beta" in token:
                    beta = float(worksheet[cell2].value)
                elif "gamma" in token:
                    gamma = float(worksheet[cell2].value)
        if avector is not None and bvector is not None and cvector is not None:
            unitCell = UnitCell(a=avector,b=bvector,c=cvector)
        elif a is not None and b is not None and c is not None and alpha is not None and beta is not None and gamma is not None:
            unitCell = UnitCell(a=a,b=b,c=c,alpha=alpha,beta=beta,gamma=gamma)
        # Process the entry type
        if "constant" in entry and "refractive" in entry:
            material = self.readConstantRefractiveIndex(sheet,worksheet,density)
        elif "constant" in entry and ("permitt" in entry or "dielec" in entry):
            material = self.readConstantPermittivity(sheet,worksheet,density)
        elif "tabulated" in entry and "refractive" in entry:
            material = self.readTabulatedRefractiveIndex(sheet,worksheet,density)
        elif "tabulated" in entry and "spec" in entry:
            material = self.readTabulatedSpectroscopy(sheet,worksheet,density)
        elif "tabulated" in entry and ("permitt" in entry or "dielec" in entry):
            material = self.readTabulatedPermittivity(sheet,worksheet,density)
        elif "lorentz" in entry and "drude" in entry:
            material = self.readLorentzDrude(sheet,worksheet,density,unitCell)
        elif "fpsq" in entry:
            material = self.readFPSQ(sheet,worksheet,density,unitCell)
        elif "sellmeier" in entry:
            material = self.readSellmeier(sheet,worksheet,density,unitCell)
        # Close the work book
        # workbook.close()
        # Add the material to the cache
        self.cache[sheet] = material
        return material

    def readConstantRefractiveIndex(self,sheet,worksheet,density):
        """Read constant refractive index from the spreadsheet.

        Parameters
        ----------
        sheet : str
            The worksheet name.
        worksheet : worksheet
            The worksheet.
        density : float
            The density of the material.

        Returns
        -------
        None

        """
        # Constant refractive index
        n = float(worksheet["C2"].value)
        k = float(worksheet["D2"].value)
        nk = complex(n, k)
        permittivity = Calculator.calculate_permittivity(nk)
        debugger.print("Constant refractive:: ",nk,permittivity,density)
        return Constant(sheet,permittivity=permittivity,density=density)

    def readConstantPermittivity(self,sheet,worksheet,density):
        """Read constant permittivity data from the spreadsheet.

        Parameters
        ----------
        sheet : str
            The worksheet name.
        worksheet : worksheet
            The worksheet instance.
        density : float
            The density of the material.

        Returns
        -------
        None

        """
        # Constant permittivity
        eps_r = float(worksheet["C2"].value)
        eps_i = float(worksheet["D2"].value)
        permittivity = complex(eps_r, eps_i)
        debugger.print("Constant permittivity:: ",permittivity,density)
        return Constant(sheet,permittivity=permittivity,density=density)

    def readTabulatedRefractiveIndex(self,sheet,worksheet,density):
        """Read tabulated refractive index data from the spreadsheet.

        Parameters
        ----------
        sheet : str
            The worksheet name.
        worksheet : worksheet
            The worksheet.
        density : float
            The density of the material.

        Returns
        -------
        None

        """
        # Tabulated refractive index
        permittivities = []
        vs_cm1 = []
        for a, c, d in zip(worksheet["A"][1:] ,worksheet["C"][1:] , worksheet["D"][1:]):
            if a.value is None or c.value is None or d.value is None:
                break
            try:
               v = float(a.value)
               n = float(c.value)
               k = float(d.value)
               nk = complex(n, k)
               permittivity = Calculator.calculate_permittivity(nk)
               permittivities.append(permittivity)
               vs_cm1.append(v)
            except Exception:
                print("Error in Tabulated: ",a.value,c.value,d.value)
        return Tabulated(sheet,vs_cm1,permittivities=permittivities,density=density)

    def readTabulatedSpectroscopy(self,sheet,worksheet,density):
        """Read tabulated refractive index data from the spreadsheet.

        Parameters
        ----------
        sheet : str
            The worksheet name.
        worksheet : worksheet
            The worksheet.
        density : float
            The density of the material.

        Returns
        -------
        None

        """
        # Tabulated refractive index + absorption (in cm^-1)
        permittivities = []
        vs_cm1 = []
        for a, c, d in zip(worksheet["A"][1:] ,worksheet["C"][1:] , worksheet["D"][1:]):
            if a.value is None or c.value is None or d.value is None:
                break
            try:
               v = float(a.value)
               n = float(c.value)
               a = float(d.value)
               k = a /(v * 4 * np.pi * math.log10(math.e))
               nk = complex(n, k)
               permittivity = Calculator.calculate_permittivity(nk)
               permittivities.append(permittivity)
               vs_cm1.append(v)
            except Exception:
                print("Error in Tabulated: ",a.value,c.value,d.value)
        #return permittivities, vs_cm1
        return Tabulated(sheet,vs_cm1,permittivities=permittivities,density=density)        
        


    def readTabulatedPermittivity(self,sheet,worksheet,density):
        """Read tabulated permittivity data from the spreadsheet.

        Parameters
        ----------
        sheet : str
            The worksheet name.
        worksheet : worksheet
            The worksheet.
        density : float
            The density of the material.

        Returns
        -------
        None

        """
        # Tabulated permittivity
        permittivities = []
        vs_cm1 = []
        for a, c, d in zip(worksheet["A"][1:] ,worksheet["C"][1:] , worksheet["D"][1:]):
            if a.value is None or c.value is None or d.value is None:
                break
            v = float(a.value)
            eps_r = float(c.value)
            eps_i = float(d.value)
            permittivity = complex(eps_r, eps_i)
            permittivities.append(permittivity)
            vs_cm1.append(v)
        return Tabulated(sheet,vs_cm1,permittivities=permittivities,density=density)

    def readLorentzDrude(self,sheet,worksheet,density,unitCell):
        """Read Drude-Lorentz data from the spreadsheet.

        Parameters
        ----------
        sheet : str
            The worksheet name.
        worksheet : worksheet
            The worksheet.
        density : float
            The density of the material.
        unitCell : object
            The unit cell.

        Returns
        -------
        None

        """
        # Lorentz-Drude model for permittivity
        epsilon_infinity = np.zeros( (3,3) )
        omegas = [[], [], []]
        strengths = [[], [], []]
        gammas = [[], [], []]
        for a, b, c, d, e in zip(worksheet["A"][1:] ,worksheet["B"][1:] , worksheet["C"][1:], worksheet["D"][1:], worksheet["E"][1:]) :
            try:
                if a.value is not None:
                    direction = a.value
                index = ["xx","yy","zz"].index(direction)
                if b.value is not None:
                    epsilon_infinity[[index],[index]] = float(b.value)
                if c.value is not None:
                    omegas[index].append(float(c.value))
                if d.value is not None:
                    strengths[index].append(float(d.value))
                if e.value is not None:
                    gammas[index].append(float(e.value))
            except Exception:
                print("Error in Lorentz-Drude: ",a.value,b.value,c.value,d.value,e.value)
                return None
        return DrudeLorentz(sheet,epsilon_infinity,omegas,strengths,gammas,density=density,cell=unitCell)

    def readFPSQ(self,sheet,worksheet,density,unitCell):
        """Read FPSQ data from the spreadsheet.

        Parameters
        ----------
        sheet : str
            The worksheet name.
        worksheet : worksheet
            The actual worksheet object.
        density : float
            The density of the material.
        unitCell : object
            The unit cell.

        Returns
        -------
        None

        """
        # FPSQ model for permittivity
        epsilon_infinity = np.zeros( (3,3) )
        omega_tos = [[], [], []]
        gamma_tos = [[], [], []]
        omega_los = [[], [], []]
        gamma_los = [[], [], []]
        for a, b, c, d, e, f in zip(worksheet["A"][1:] ,worksheet["B"][1:] , worksheet["C"][1:], worksheet["D"][1:], worksheet["E"][1:], worksheet["F"][1:]) :
            try:
                if a.value is not None:
                    direction = a.value
                index = ["xx","yy","zz"].index(direction)
                if b.value is not None:
                    epsilon_infinity[[index],[index]] = float(b.value)
                if c.value is not None:
                    omega_tos[index].append(float(c.value))
                if d.value is not None:
                    gamma_tos[index].append(float(d.value))
                if e.value is not None:
                    omega_los[index].append(float(e.value))
                if f.value is not None:
                    gamma_los[index].append(float(f.value))
            except Exception:
                print("Error in FPSQ: ",a.value,b.value,c.value,d.value,e.value,f.value)
                return None
        return FPSQ(sheet,epsilon_infinity,omega_tos,gamma_tos,omega_los,gamma_los,density=density,cell=unitCell)

    def readSellmeier(self,sheet,worksheet,density,unitCell):
        """Read Sellmeier data from the spreadsheet.

        Parameters
        ----------
        sheet : str
            The worksheet name.
        worksheet : worksheet
            The worksheet.
        density : float
            The density of the material.
        unitCell : object
            The unit cell.

        Returns
        -------
        None

        """
        # Sellmeier model for refractive index
        Bs = []
        Cs = []
        for b, c in zip(worksheet["A"][1:] , worksheet["B"][1:] ) :
            try:
                if b.value is not None:
                    Bs.append(float(b.value))
                if c.value is not None:
                    Cs.append(float(c.value))
            except Exception:
                print("Error in Sellmeier: ",b.value,c.value)
                return None
        return Sellmeier(sheet,Bs,Cs,density=density,cell=unitCell)
        
class Material:
    """A class for representing materials with properties like name, density, permittivity, and unit cell.

    The Material class also contains the permittivity object for the material.  
    The permittivity object is an instance of :class:`~PDielec.DielectricFunction.DielectricFunction` or one its children.
    The permittivity object is responsible for calculating the permittivity at the given frequency.
    The subclasses which inherit from the Material class are: Constant, External, DrudeLorentz, FPSQ, Sellmeier, and Tabulated.
    Each subclass has an initialisation routine which instantiates the permittivity object of the appropriate type for the material.
    The relationship between the Material subclass and the DielectricFunction subclass of the permittivity object is shown below.

    +---------------------------+-------------------------------------------------------------+
    + Material subclass         + DielectricFunction subclass                                 +
    +===========================+=============================================================+
    + :class:`Constant`         + :class:`~PDielec.DielectricFunction.Constant`               +
    +---------------------------+-------------------------------------------------------------+
    + :class:`External`         + This class is passed a permittivity object which has been   +
    +                           + defined externally                                          +
    +---------------------------+-------------------------------------------------------------+
    + :class:`DrudeLorentz`     + :class:`~PDielec.DielectricFunction.DrudeLorentz`           +
    +---------------------------+-------------------------------------------------------------+
    + :class:`FPSQ`             + :class:`~PDielec.DielectricFunction.FPSQ`                   +
    +---------------------------+-------------------------------------------------------------+
    + :class:`Sellmeier`        + :class:`~PDielec.DielectricFunction.Sellmeier`              +
    +---------------------------+-------------------------------------------------------------+
    + :class:`Tabulated`        +  - :class:`~PDielec.DielectricFunction.TabulateScalar`      +
    +                           +  - :class:`~PDielec.DielectricFunction.Tabulate3`           +
    +                           +  - :class:`~PDielec.DielectricFunction.Tabulate3`           +
    +                           +  - :class:`~PDielec.DielectricFunction.Tabulate6`           +
    +---------------------------+-------------------------------------------------------------+

    Parameters
    ----------
    name : str
        The name of the material.
    density : float, optional
        The density of the material. If not provided and a cell is given, it will be calculated based on the cell.
    permittivityObject : :class:`~PDielec.DielectricFunction.DielectricFunction`, optional
        An object representing the dielectric function of the material. This is intended to be passed by classes that inherit from Material, and it should contain methods for calculating scalar/tensor permittivity. (see :class:`~PDielec.DielectricFunction.DielectricFunction` and its sub-classes)
    cell : :class:`~PDielec.UnitCell.UnitCell`, optional
        An object representing the unit cell of the material. If provided without a density, the density will be calculated from this cell. (See :class:`~PDielec.UnitCell.UnitCell`)

    Attributes
    ----------
    density : float
        The density of the material, which may be calculated based on the cell if not provided initially.
    cell : Cell or None
        The unit cell of the material if provided.
    name : str
        The name of the material.
    type : str
        A string indicating the type of the object. Defaults to 'Base Class' for the base Material class.
    permittivityObject : DielectricFunction or None
        An object to handle the permittivity calculations for the material.

    Methods
    -------
    getName()
        Returns the name of the material.
    getInformation()
        Returns information about the material, including its type and, if applicable, its permittivity frequency range.
    getSigmas()
        If the material has a lorentzian dielectric this routine returns the sigma parameters
    setSigmas()
        If the material has a lorentzian dielectric this routine sets the sigma parameters
    getFrequencies()
        If the material has a lorentzian dielectric this routine returns the frequencies
    setFrequencies()
        If the material has a lorentzian dielectric this routine sets the frequencies
    getOscillatorStrengths()
        If the material has a lorentzian dielectric this routine returns the oscillator strengths
    setOscillatorStrengths()
        If the material has a lorentzian dielectric this routine sets the oscillator strengths
    print()
        Prints information about the material, such as its name, density, type, and permittivity details.
    isScalar()
        Checks and returns True if the material’s permittivity is scalar.
    isTensor()
        Checks and returns True if the material’s permittivity is tensor.
    getPermittivityObject()
        Returns the permittivityObject of the material.
    getPermittivityFunction()
        Returns the permittivity function from the permittivityObject.
    getDensity()
        Returns the density of the material.
    setCell(cell)
        Sets the cell of the material and updates the density if it was initially None.
    getCell()
        Returns the cell of the material.
    setDensity(value)
        Sets the density of the material.
    setEpsilonInfinity(eps)
        Sets the epsilon infinity of the material
    setPermittivityObject(permittivityObject)
        Sets the permittivityObject for the material.

    """

    def __init__(self, name, density=None, permittivityObject=None, cell=None):
        """Initialise a material with the following parameters.

        Parameters
        ----------
        name : str
            The name of the material.
        density : float, optional
            The density of the material. If not provided and a cell is given, it will be calculated based on the cell.
        permittivityObject : DielectricFunction, optional
            An object representing the dielectric function of the material. This is intended to be passed by classes that inherit from Material, and it should contain methods for calculating scalar/tensor permittivity.
        cell : unitCell, optional
            An object representing the unit cell of the material. If provided without a density, the density will be calculated from this cell.

        Notes
        -----
        The material object is created from the name, density, and unit cell. The permittivity object is specifically created by the children of Material, indicating it's a derived property or capability not initialized directly by the Material's constructor but through some other process or method within the child classes.

        """
        self.density            = density
        self.cell               = cell
        self.name               = name
        self.type               = "Base Class"
        self.permittivityObject = permittivityObject
        if self.density is None and self.cell is not None:
            self.density = self.cell.getDensity("cm")

    def getName(self):
        """Get the name attribute of the object.

        Parameters
        ----------
        None

        Returns
        -------
        str
            The name attribute of the object.

        """        
        return self.name

    def getInformation(self):
        """Return information about the material.

        Parameters
        ----------
        None

        Returns
        -------
        str
            A description of the material.

        """
        result = self.type
        if "Tabulate" in self.type:
            low = self.permittivityObject.getLowestFrequency()
            high = self.permittivityObject.getHighestFrequency()
            result += f" freq range {low:.0f}-{high:.0f}" # + ' value at 0 {}'.format(self.permittivityObject.function()(0))
        # result += ' value at 0 {}'.format(self.permittivityObject.function()(0))
        return result

    def print(self):
        """Print information about the material.

        Parameters
        ----------
        None

        Returns
        -------
        None

        """
        print("Material name:",self.name)
        print("Material density:",self.density)
        print("Material type:",self.type)
        print("Material is scalar?:",self.isScalar())
        print("Material is tensor?:",self.isTensor())
        print("Material permittivity:",self.getInformation())
        if self.cell is not None:
            print("Material unit cell")
            self.cell.print()
        return

    def isScalar(self):
        """Return true if the material returns a scalar permittivity.

        Parameters
        ----------
        None

        Returns
        -------
        bool
            True if the material returns a scalar permittivity, False otherwise.

        """
        return self.permittivityObject.isScalar()

    def isTensor(self):
        """Return true if the material returns a tensor permittivity.

        Parameters
        ----------
        None

        Returns
        -------
        bool
            True if the material returns a tensor permittivity, False otherwise.

        """
        return self.permittivityObject.isTensor()

    def setPermittivityObject(self,permittivityObject):
        """Set the permittivity object.

        Parameters
        ----------
        permittivityObject : a permittivity object (dielectric function object)
            The permittivity object is used to calculate the permittivity of the material

        Returns
        -------
        None

        """
        self.permittivityObject = permittivityObject
        return

    def getPermittivityObject(self):
        """Return the permittivity object.

        Parameters
        ----------
        None

        Returns
        -------
        permittivityObject
            Return the permittivity object (dielectric function object)

        """
        return self.permittivityObject

    def getPermittivityFunction(self):
        """Return the permittivity function.

        Parameters
        ----------
        None

        Returns
        -------
        permittivityObjectFunction
            Return the permittivity object function

        """
        return self.permittivityObject.function()

    def setFrequencies(self,frequencies):
        """Set the frequencies for a Lorentzian permittivity.

        Parameters
        ----------
        frequencies : 1D array of floats
            The frequencies for a Drude-Lorentzian permittivity in cm-1

        Returns
        -------
        None

        """
        self.permittivityObject.setFrequencies(frequencies)
        return 

    def getFrequencies(self):
        """Get the frequencies for a Lorentzian permittivity.

        Parameters
        ----------
        None

        Returns
        -------
        1d array of floats
            Returns the frequencies for a Lorentzian function in cm-1

        """
        return self.permittivityObject.getFrequencies()

    def setOscillatorStrengths(self,strengths):
        """Set the oscillator strengths for a Lorentzian permittivity.

        Parameters
        ----------
        strengths : a 3x3 array of floats for each frequency
            The oscillator strengths for a Lorentzian permittivity function in cm-1

        Returns
        -------
        None

        """
        self.permittivityObject.setOscillatorStrengths(strengths)
        return 

    def getOscillatorStrengths(self):
        """Get the oscillator strengths for a Lorentzian permittivity.

        The oscillator strength of each transition is a 3x3 matrix

        Parameters
        ----------
        None

        Returns
        -------
        list of 3x3 array of floats
            Returns the oscillator strengths for a Lorentzian permittivity function in cm-1

        """
        return self.permittivityObject.getOscillatorStrengths()


    def setSigmas(self,sigmas):
        """Set the sigma parameters for a Lorentzian permittivity.

        Parameters
        ----------
        sigmas : 1D array of floats
            The sigma parameters for a Lorentzian permittivity function in cm-1

        Returns
        -------
        None

        """
        self.permittivityObject.setSigmas(sigmas)
        return 

    def getSigmas(self):
        """Get the sigma parameters for a Lorentzian permittivity.

        Parameters
        ----------
        None

        Returns
        -------
        1d array of floats
            Returns the sigma parameters for a Lorentz permittivity function in cm-1

        """
        return self.permittivityObject.getSigmas()


    def setDensity(self, value):
        """Set the density.

        Parameters
        ----------
        value : float
            The value of the density

        Returns
        -------
        None

        """
        self.density = value
        return

    def getDensity(self):
        """Return the density.

        Parameters
        ----------
        None

        Returns
        -------
        self.density

        """
        return self.density

    def setCell(self, cell):
        """Set the unit cell.

        Parameters
        ----------
        cell : a unit cell
            Set the unit cell of the material

        Returns
        -------
        None

        """
        self.cell = cell
        if self.density is None and self.cell is not None:
            self.density = self.cell.calculate_density()
        return

    def getCell(self):
        """Return the cell.

        Parameters
        ----------
        None

        Returns
        -------
        self.cell

        """
        return self.cell

class Constant(Material):
    """A class representing a material with constant scalar permittivity, inheriting from the `Material` class.

    Attributes
    ----------
    type : str
        The type of material, set to 'Constant permittivity'.

    Methods
    -------
    Inherits methods from the `Material` class.

    """

    def __init__(self, name, permittivity=None, density=None, cell=None):
        """Create an instance of a material with a constant scalar permittivity.

        Permittivity is the value of the permittivity and can be complex.

        Parameters
        ----------
        name : str
            The name of the material.
        permittivity : complex
            The permittivity value. It can be a complex number.
        density : float
            The density of the material in g/ml.
        cell : unitCell
            The unit cell.

        """
        super().__init__(name, density=density, permittivityObject=DielectricFunction.ConstantScalar(permittivity), cell=cell)
        self.type = "Constant permittivity"


class External(Material):
    """A class for representing materials with externally specified permittivity.

    This class inherits from the `Material` class and is used to define materials
    where the permittivity is specified externally, rather than calculated or predefined.
    Permittivity can be a complex value indicating both the real and imaginary parts.
    """

    def __init__(self, name, permittivityObject=None, density=None, cell=None):
        """Create an instance of a material which has the permittivity object specified externally.

        The permittivity can be a complex number.

        Parameters
        ----------
        name : str
            The name of the material.
        permittivityObject : complex
            The permittivity value, which can be a complex number.
        density : float
            The density of the material in grams per milliliter (g/ml).
        cell : unitCell
            The unit cell of the material.

        Returns
        -------
        object
            An instance of the material with the specified permittivity object.

        """
        super().__init__(name, density=density, permittivityObject=permittivityObject, cell=cell)
        self.type = "External permittivity"


class DrudeLorentz(Material):
    """A subclass representing a material with a Lorentz-Drude model permittivity.

    Parameters
    ----------
    name : str
        The name of the material.
    epsinf : array_like
        Epsilon infinity, either a 3x3 list or a 3x3 array representing the static dielectric constant.
    omegas : list
        The transverse optical (TO) frequencies.
    strengths : list
        The oscillator strengths for each resonance.
    gammas : list
        The damping (or broadening) factors for each resonance.
    density : float, optional
        The density of the material in grams per milliliter (g/ml). Default is None.
    cell : unitCell, optional
        The unit cell of the material. Default is None.

    Notes
    -----
    The Drude-Lorentz model is used to calculate the permittivity of the material
    by considering the contributions from both free electrons (Drude) and bound electrons
    (Lorentz). This class requires specifying the infinite frequency dielectric constant (`epsinf`),
    the transverse optical frequencies (`omegas`), the oscillator strengths (`strengths`),
    and the damping factors (`gammas`) for each resonance in the material.

    Examples
    --------
    >>> drude_lorentz_material = DrudeLorentz("Gold", [[1, 0, 0], [0, 1, 0], [0, 0, 1]],
    ...                                       [0.5, 1.0], [1.0, 2.0], [0.2, 0.1],
    ...                                       density=19.3)
    This represents a Drude-Lorentz material with the name "Gold", an isotropic epsilon infinity,
    two resonances with specified frequencies, strengths, and damping factors, and 
    a density of 19.3 g/ml.

    """

    def __init__(self, name,epsinf,omegas,strengths,gammas,density=None,cell=None):
        """Create an instance of a material with a Lorentz Drude model permittivity.

        Parameters
        ----------
        name : str
            The name of the material.
        epsinf : list or array
            Epsilon infinity either a 3x3 list or a 3x3 array.
        omegas : list
            The Transverse Optical (TO) frequencies.
        strengths : list
            The absorption strengths.
        gammas : list
            The absorption widths.
        density : float
            The density of the material in g/ml.
        cell : unitCell
            The unit cell of the material.

        """
        epsilon_infinity = np.array(epsinf)
        permittivityObject = DielectricFunction.DrudeLorentz( omegas, strengths, gammas)
        permittivityObject.setEpsilonInfinity(epsilon_infinity)
        super().__init__(name, density=density, permittivityObject=permittivityObject,cell=cell)
        self.type = "Drude-Lorentz"

class FPSQ(Material):
    """Class representing a material with a FPSQ model for permittivity.

    Parameters
    ----------
    name : str
        The name of the material.
    epsinf : array_like
        Epsilon infinity (eps0), either a 3x3 list or a 3x3 array representing the dielectric constant at infinite frequency.
    omega_tos : list
        The transverse optical (TO) frequencies.
    gamma_tos : list
        The TO absorption widths.
    omega_los : list
        The longitudinal optical (LO) frequencies.
    gamma_los : list
        The LO absorption widths.
    density : float, optional
        The density of the material in grams per milliliter (g/ml).
    cell : unitCell, optional
        The unit cell of the material.

    Notes
    -----
    The FPSQ (fitted phonon simple quantum) model is used to describe the permittivity of the material. This model is based on the harmonic oscillator model and describes the permittivity as a function of frequency.

    """

    def __init__(self, name,epsinf,omega_tos,gamma_tos,omega_los,gamma_los,density=None,cell=None):
        """Create an instance of a material with an FPSQ model permittivity.

        Parameters
        ----------
        name : str
            The name of the material.
        epsinf : list or ndarray
            Epsilon infinity (ε∞), either a 3x3 list or a 3x3 array.
        omega_tos : list
            The transverse optical (TO) frequencies.
        gamma_tos : list
            The TO absorption widths.
        omega_los : list
            The longitudinal optical (LO) frequencies.
        gamma_los : list
            The LO absorption widths.
        density : float
            Density in g/ml.
        cell : unitCell
            The unit cell.

        """
        epsilon_infinity = np.array(epsinf)
        permittivityObject = DielectricFunction.FPSQ( omega_tos, gamma_tos, omega_los, gamma_los)
        permittivityObject.setEpsilonInfinity(epsilon_infinity)
        super().__init__(name, density=density, permittivityObject=permittivityObject,cell=cell)
        self.type = "FPSQ"

class Sellmeier(Material):
    """A class to define materials using the Sellmeier model for permittivity.

    Parameters
    ----------
    name : str
        The name of the material.
    Bs : list or array_like
        The B parameters (coefficients) in the Sellmeier equation.
    Cs : list or array_like
        The C parameters (coefficients) in the Sellmeier equation.
    density : float, optional
        The density of the material in g/ml. Default is None.
    cell : unitCell, optional
        The unit cell of the material. Default is None.

    Attributes
    ----------
    type : str
        The type of the material, which is 'Sellmeier' for instances of this class.

    Methods
    -------
    __init__(self, name, Bs, Cs, density=None, cell=None)
        Initializes a Sellmeier material with specified parameters.

    """

    def __init__(self, name,Bs,Cs,density=None,cell=None):
        """Create an instance of a material with a Sellmeier model permittivity.

        Permittivity is the value of the permittivity and should be real for the Sellmeier model.
        The required parameters are:

        Parameters
        ----------
        name : str
            The name of the material.
        Bs : list
            The B parameters of the Sellmeier equation.
        Cs : list
            The C parameters of the Sellmeier equation.
        density : float
            Density in g/ml.
        cell : unitCell
            The unit cell.

        """
        permittivityObject = DielectricFunction.Sellmeier( Bs, Cs)
        super().__init__(name, density=density, permittivityObject=permittivityObject,cell=cell)
        self.type = "Sellmeier"


class Tabulated(Material):
    """A class for materials with tabulated permittivities.

    Parameters
    ----------
    name : str
        The name of the material.
    vs_cm1 : list or None, optional
        The list of tabulated frequencies in cm-1. Defaults to None.
    permittivities : array-like or None, optional
        The permittivities, either as a single vector (n,) or a tensor (3,n) or (6,n) for more complex materials.
        Defaults to None.
    density : float or None, optional
        The density of the material in g/ml. Defaults to None.
    cell : unitCell or None, optional
        The unit cell of the material. Defaults to None.

    Notes
    -----
    - This class is designed to handle materials with a constant permittivity as well as those
      requiring more complex permittivity tensors.
    - The permittivity can be defined using either a scalar for simple materials or tensors for materials
      that require a support matrix.
    - The constructor converts the input lists of frequencies (`vs_cm1`) and permittivities into numpy arrays,
      and then generates the appropriate permittivity object depending on the complexity of the material's permittivities.

    Examples
    --------
    >>> material1 = Tabulated("Quartz", vs_cm1=[500, 1000, 1500], permittivities=[2.1, 2.3, 2.5], density=2.65)
    >>> material2 = Tabulated("Synthetic", vs_cm1=[200, 400, 600], permittivities=[[2.1, 2.3, 2.5], [2.4, 2.6, 2.8], [3.0, 3.2, 3.4]], density=1.5)

    """

    def __init__(self, name, vs_cm1=None, permittivities=None, density=None, cell=None):
        """Create an instance of a material with a constant permittivity. Permittivity is the value of the permittivity and can be complex. The returned permittivityObject can generate either a scalar or a tensor. For defining a support matrix material, a scalar is used.

        Parameters
        ----------
        name : str
            The name of the material.
        vs_cm1 : list
            The list of tabulated frequencies in cm-1.
        permittivities : array_like
            The permittivities, either a single (n) vector or a (3,n) vector.
        density : float
            Density in g/ml.
        cell : unitCell
            The unit cell.

        """
        vs = np.array(vs_cm1)
        eps = np.array(permittivities)
        if len(np.shape(eps)) == 2:
            m,n = np.shape(eps)
            if m == 3:
                permittivityObject = DielectricFunction.Tabulate3(vs,eps[0], eps[1], eps[2])
            elif m== 6:
                permittivityObject = DielectricFunction.Tabulate6(vs,eps[0], eps[1], eps[2], eps[3], eps[4], eps[5])
            else:
                print("Error in Tabulated, shape of parameters is wrong")
        else:
            permittivityObject = DielectricFunction.TabulateScalar(vs,eps)
        super().__init__(name, density=density, permittivityObject=permittivityObject,cell=cell)
        self.type = "Tabulated permittivity"

    def setEpsilonInfinity(self,eps):
        """Set the value of epsilon infinity for the material.

        Parameters
        ----------
        eps : float or 3x3 np array
            The epsilon infinity tensor.  If a single float then an isotropic 3x3 np array is created

        Returns
        -------
        None

        """
        eps = eps*np.eye(3) if isinstance(eps,float) else np.array(eps)
        self.permittivityObject.setEpsilonInfinity(eps)

