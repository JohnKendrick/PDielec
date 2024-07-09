#!/usr/bin/env python
#
# Copyright 2024 John Kendrick & Andrew Burnett
#
# This file is part of PDielec
#
# This program is free software; you can redistribute it and/or modify
# it under the terms of the MIT License
#
# This program is distributed in the hope that it will be useful,
# but WITHOUT ANY WARRANTY; without even the implied warranty of
# MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.
#
# You should have received a copy of the MIT License along with this program, if not see https://opensource.org/licenses/MIT
#
"""The checkexcel package reads in two excel files and check that the float contents are the same.

It is used by the :mod:`~PDielec.pdmake` command to check the validity of the reference calculations in the test suite.

"""
import sys

from openpyxl import load_workbook
from termcolor import colored

global threshold

def main():

    # Start processing the directories
    """Compare two Excel files for any significant numerical changes and report discrepancies.

    This script takes in two Excel files as arguments and optionally,
    a threshold for numerical comparison and a flag for full sheet comparison.

    Full comparison includes 'Settings' and 'Scenarios' sheets by default excluded.

    Parameters
    ----------
    None directly to the function. Parses command line arguments:
    - Two file paths.
    - Optional `-thresh` followed by a float value for the threshold of numerical difference (default is 1.0E-3).
    - Optional `-f` flag for forcing a full comparison including typically excluded sheets.

    Returns
    -------
    tuple
        A tuple containing the following elements:
        - Number of errors (`nerrors`): int
        - Row of the last error or discrepancy (`row`): int
        - Column of the last error or discrepancy (`col`): int
        - Sheet name where the last error or discrepancy was found (`sheet`): str
        - First file name (`file1`): str
        - Value from the first file associated with the last discrepancy (`value1`): float or str
        - Second file name (`file2`): str
        - Value from the second file associated with the last discrepancy (`value2`): float or str
        - Maximum percentage error found during the comparison (`max_percentage_error`): float

    Raises
    ------
    Depends on the Excel library used, typically:
    - `openpyxl.utils.exceptions.InvalidFileException` if an invalid file or path is provided.
    - Other exceptions related to file I/O or processing might be raised depending on the content and structure of the input Excel files.

    Notes
    -----
    - The script will immediately exit with usage instructions if less than two file paths are provided.
    - Numerical difference is calculated only for numeric data. For text data, a simplified equivalence check is done.
    - The script output includes printing to standard error for usage, errors, or status, with optional ANSI color highlighting.

    """    
    if len(sys.argv) <= 1 :
        print("checkexcel file file2 [-thresh 1.0E-3] [-f]", file=sys.stderr)
        print("         Compare the two excel files for any significant changes", file=sys.stderr)
        print("         Numbers f1 and f2 in each file are compared", file=sys.stderr)
        print("         an error is flagged if 2*abs(f1-f2)/(f1+f2+2) > threshold", file=sys.stderr)
        print("         The default value for threshold is 1.0E-3 ", file=sys.stderr)
        print("         -f forces a full comparison of the sheets ", file=sys.stderr)
        print("            by default Settings and Scenarios are not included    ", file=sys.stderr)
        sys.exit()

    threshold = 1.0E-3
    tokens = sys.argv[1:]
    ntokens = len(tokens)-1
    itoken = -1
    files = []
    full = False
    while itoken < ntokens:
        itoken += 1
        token = tokens[itoken]
        if token == "-thresh":
            itoken +=1
            threshold = float(tokens[itoken])
        if token == "-f":
            full = True
        else:
            files.append(tokens[itoken])
        # end if
    # end while

    file1 = files[0]
    file2 = files[1]

    #
    # Open the work books
    #
    wb1 = load_workbook(filename=file1, read_only=True)
    wb2 = load_workbook(filename=file2, read_only=True)
    #
    # Initialise variables
    #
    error = None
    nerrors = 0
    max_percentage_error = 0.0
    row = 0
    col = 0
    sheet = ""
    file1 = "" 
    value1 = 0
    file2 = "" 
    value2 = 0
    #
    # Loop over sheets
    #
    sheets = ["Powder Molar Absorption (cells)","Powder Absorption","Powder Real Permittivity","Powder Imaginary Permittivity", "Powder ATR Reflectance", "Analysis","Crystal R_p","Crystal R_s","Crystal T_p","Crystal T_s","Real Crystal Permittivity","Imag Crystal Permittivity"]
    if full:
        sheets.append("Main")
        sheets.append("Settings")
        sheets.append("Scenarios")
    for sheet in sheets:
        if sheet not in wb1 :
            continue
        if sheet not in wb2 :
            continue
        print("Checking sheet ",sheet)
        ws1 = wb1[sheet]
        ws2 = wb2[sheet]
        max_rows1 = ws1.max_row
        max_rows2 = ws2.max_row
        max_columns1 = ws1.max_column
        max_columns2 = ws2.max_column
        if max_rows1 != max_rows2:
            print("Error - the number rows in the sheet are not the same",sheet,max_rows1,max_rows2)
            nerrors += 1
            continue
        if max_columns1 != max_columns2:
            print("Error - the number columns in the sheet are not the same",sheet,max_columns1,max_columns2)
            nerrors += 1
            continue
        #
        # Loop over rows
        #
        for row_index,(row1,row2) in enumerate(zip(ws1.rows, ws2.rows)):
            #
            # Loop over cells
            #
            for col_index,(cell1,cell2) in enumerate(zip(row1,row2)):
                value1 = cell1.value
                value2 = cell2.value
                if cell1 is None and cell2 is None:
                    pass
                elif cell1 is None or cell2 is None:
                    nerrors += 1
                    percentage_error = 0.0
                    error = (sheet, row_index, col_index, value1, value2, percentage_error)
                elif value1 != value2:
                    if value1 is not None and value2 is not None and cell1.data_type == "n" and cell2.data_type == "n":
                        #
                        # Flag an error which is numeric
                        #
                        percentage_error = 100.0*abs(2.0*(value1 - value2)/(abs(value1) + abs(value2)+2))
                        if percentage_error > 100.*threshold:
                            nerrors += 1
                            if percentage_error > max_percentage_error:
                                max_percentage_error = percentage_error
                                error = (sheet, row_index, col_index, value1, value2, percentage_error)
                            # if percentage error
                        elif percentage_error > max_percentage_error:
                            max_percentage_error = percentage_error
                        # if percentage error > threshold
                    else:
                        #
                        # This is a non-numeric error
                        # Remove any back or forward slashes to avoid problems with filenames in linux/windows
                        #
                        if isinstance(value1,str):
                            value1 = value1.replace("\\","")
                            value1 = value1.replace("/","")
                        if isinstance(value2,str):
                            value2 = value2.replace("\\","")
                            value2 = value2.replace("/","")
                        if value1 != value2:
                            nerrors += 1
                            error = (sheet, row_index, col_index, value1, value2, 0.0)
                        # if cell1.data_type
                    # if value1 != value2
                # if cell1 is none
            # for cell1, cell2
        # for rowq, row
    # for sheet
    if error is not None:
        sheet,row,col,value1,value2,max_percentage_error = error
        print("  "+colored("ERRORS:","red")+f"({nerrors}) LARGEST ON ROW,COL {row},{col} OF SHEET {sheet}, {file1}({value1}) and {file2}({value2}) -- max %error={max_percentage_error}")
    elif nerrors > 0:
        print("  "+colored("ERRORS:","red")+f"({nerrors}) Dimensions of spreadsheet were wrong                                    ")
    else:
        print("  "+colored("OK:","blue")+f" {file1} = {file2} -- max %error={max_percentage_error}")
    return nerrors, row,col,sheet,file1,value1,file2,value2,max_percentage_error

if __name__=="__main__":
    main()
